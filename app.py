# -*- coding: utf-8 -*-
import sys
import os

# Handle stdout encoding (only if stdout is available - not in noconsole mode)
if sys.stdout is not None:
    sys.stdout.reconfigure(encoding='utf-8')

# Helper function to get resource path (works with PyInstaller)
def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timedelta
import json
import pandas as pd
from io import BytesIO
import requests
import pickle
import threading
import tempfile
from config import Config
app = Flask(__name__)
app.config.from_object(Config)


# Google Sheets configuration
SCOPE = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

# Global cache for Google Sheets clients to avoid recreating them
_google_sheets_clients = {}
_client_lock = threading.Lock()

def get_google_sheets_client(data_type=None):
    """Initialize Google Sheets client"""
    try:
        # Determine which credentials file to use based on data type
        if data_type in ['egypt_match', 'egypt_lineup']:
            creds_file = get_resource_path('credentials/egyptnationalteam.json')
        else:
            creds_file = get_resource_path(app.config['GOOGLE_CREDENTIALS_FILE'])
        
        if os.path.exists(creds_file):
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        else:
            # Try to get from specific environment variable based on data type
            if data_type == 'ahly_match' and app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']:
                creds_json = app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']
            elif data_type == 'ahly_goals_assists' and app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']:
                # Use same credentials as ahly_match for Goals & Assists
                creds_json = app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']
            elif data_type == 'ahly_gks' and app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']:
                # Use same credentials as ahly_match for GKS
                creds_json = app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']
            elif data_type == 'ahly_howpenmissed' and app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']:
                # Use same credentials as ahly_match for HOWPENMISSED
                creds_json = app.config['GOOGLE_CREDENTIALS_JSON_AHLY_MATCH']
            elif data_type in ['egypt_match', 'egypt_lineup'] and app.config['GOOGLE_CREDENTIALS_JSON_EGYPT_TEAMS']:
                # Use Egypt Teams credentials for Egypt Match and Lineup
                creds_json = app.config['GOOGLE_CREDENTIALS_JSON_EGYPT_TEAMS']
            else:
                # Fallback to general credentials
                creds_json = app.config['GOOGLE_CREDENTIALS_JSON']
            
            if creds_json:
                creds_info = json.loads(creds_json)
                creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
            else:
                raise FileNotFoundError("No credentials found")
        
        return gspread.authorize(creds)
    except Exception as e:
        print(f"Error initializing Google Sheets client: {e}")
        print(f"Credentials file exists: {os.path.exists(creds_file)}")
        print(f"Environment variable exists: {bool(app.config['GOOGLE_CREDENTIALS_JSON'])}")
        print(f"Data type: {data_type}")
        return None

def save_to_sheets(data_type, data):
    """Save data to appropriate Google Sheet"""
    try:
        print(f"Attempting to save data_type: {data_type}")
        
        # Check if ahly_pks is not supported for Google Sheets
        if data_type == 'ahly_pks':
            return False, "فشل في الحفظ"
        
        client = get_google_sheets_client(data_type)
        if not client:
            return False, "فشل في الحفظ"
        
        sheet_id = app.config['SHEET_IDS'][data_type]
        worksheet_name = app.config['WORKSHEET_NAMES'][data_type]
        
        print(f"Sheet ID: {sheet_id}")
        print(f"Worksheet name: {worksheet_name}")
        
        # Open the spreadsheet
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get the worksheet, create if it doesn't exist
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)
            # Add headers only for new worksheet
            headers = get_headers_for_type(data_type)
            worksheet.append_row(headers)
        
        # Check if worksheet is empty, if so add headers
        all_values = worksheet.get_all_values()
        if not all_values:
            # Worksheet is empty, add headers
            headers = get_headers_for_type(data_type)
            worksheet.append_row(headers)
            # Refresh all_values after adding headers
            all_values = worksheet.get_all_values()
        
        # Find the last row with data
        last_row_with_data = len(all_values)
        
        # Calculate the row number where we'll start writing (leave one blank row)
        if last_row_with_data > 0:
            # If there's data (including headers), leave a blank row and write in the next one
            next_row_number = last_row_with_data + 2  # +1 for blank row, +1 for next row
        else:
            # If completely empty, start from row 1
            next_row_number = 1
        
        # Prepare data row
        if data_type == 'ahly_lineup':
            players_data, match_date, match_id = prepare_data_row(data_type, data)
            # Save each player as a separate row
            current_row = next_row_number
            for player in players_data:
                if player['name']:  # Only save if player name is provided
                    row_data = [
                        match_date,
                        match_id,
                        player['minmat'],
                        player['name'],
                        player['status'],
                        player['playerout'],
                        player['minout'],
                        player['mintotal']
                    ]
                    worksheet.insert_row(row_data, current_row)
                    current_row += 1
        elif data_type == 'egypt_lineup':
            players_data, match_date, match_id = prepare_data_row(data_type, data)
            # Save each player as a separate row (without match_date)
            current_row = next_row_number
            for player in players_data:
                if player['name']:  # Only save if player name is provided
                    row_data = [
                        match_id,
                        player['minmat'],
                        player['name'],
                        player['status'],
                        player['playerout'],
                        player['minout'],
                        player['mintotal']
                    ]
                    worksheet.insert_row(row_data, current_row)
                    current_row += 1
        elif data_type == 'ahly_match':
            # Save match data first (directly after last row, no blank row for MATCHDETAILS)
            row_data = prepare_data_row(data_type, data)
            # Override next_row_number for ahly_match to write directly after last row
            ahly_match_row = last_row_with_data + 1  # No blank row for MATCHDETAILS
            worksheet.insert_row(row_data, ahly_match_row)
            
            # Save Goals & Assists data to PLAYERDETAILS sheet
            goals_success = save_goals_assists_to_sheets(data)
            if not goals_success:
                print("Warning: Failed to save Goals & Assists data")
            
            # Save GKS data to GKDETAILS sheet
            gks_success = save_gks_to_sheets(data)
            if not gks_success:
                print("Warning: Failed to save GKS data")
            
            # Save HOWPENMISSED data to HOWPENMISSED sheet
            howpenmissed_success = save_howpenmissed_to_sheets(data)
            if not howpenmissed_success:
                print("Warning: Failed to save HOWPENMISSED data")
        elif data_type == 'egypt_match':
            # Save Egypt Match data (directly after last row, no blank row)
            row_data = prepare_data_row(data_type, data)
            egypt_match_row = last_row_with_data + 1  # No blank row for Egypt Match
            worksheet.insert_row(row_data, egypt_match_row)
        else:
            row_data = prepare_data_row(data_type, data)
            # Insert data to worksheet
            worksheet.insert_row(row_data, next_row_number)
        
        return True, "تم الحفظ"
        
    except Exception as e:
        return False, "فشل في الحفظ"

def save_goals_assists_to_sheets(data):
    """Save Goals & Assists data to PLAYERDETAILS sheet"""
    try:
        # Get Google Sheets client for Goals & Assists
        client = get_google_sheets_client('ahly_goals_assists')
        if not client:
            print("Google Sheets client not available for Goals & Assists")
            return False
        
        sheet_id = app.config['SHEET_IDS']['ahly_goals_assists']
        worksheet_name = app.config['WORKSHEET_NAMES']['ahly_goals_assists']
        
        # Open the spreadsheet
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get the worksheet, create if it doesn't exist
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)
            # Add headers for new worksheet
            headers = get_headers_for_type('ahly_goals_assists')
            worksheet.append_row(headers)
        
        # Check if worksheet is empty, if so add headers
        all_values = worksheet.get_all_values()
        if not all_values:
            # Worksheet is empty, add headers
            headers = get_headers_for_type('ahly_goals_assists')
            worksheet.append_row(headers)
        
        # Add a blank row before new data
        worksheet.append_row([''] * len(get_headers_for_type('ahly_goals_assists')))
        
        # Extract Goals & Assists data
        match_id = data.get('match_id', '')
        goals_data = []
        
        # Find all goals & assists entries
        goals_count = 0
        while f'goals_assists_{goals_count}_player_name' in data:
            player_name = data.get(f'goals_assists_{goals_count}_player_name', '')
            team = data.get(f'goals_assists_{goals_count}_team', '')
            ga = data.get(f'goals_assists_{goals_count}_ga', '')
            goal_type = data.get(f'goals_assists_{goals_count}_type', '')
            minute = data.get(f'goals_assists_{goals_count}_minute', '')
            
            # Only save if player name is provided
            if player_name:
                row_data = [match_id, player_name, team, ga, goal_type, minute]
                goals_data.append(row_data)
            
            goals_count += 1
        
        # Add goals data to sheet
        for goal_row in goals_data:
            worksheet.append_row(goal_row)
        
        print(f"Successfully saved {len(goals_data)} Goals & Assists entries to PLAYERDETAILS")
        return True
        
    except Exception as e:
        print(f"Error saving Goals & Assists data: {e}")
        return False

def save_gks_to_sheets(data):
    """Save GKS data to GKDETAILS sheet"""
    try:
        # Get Google Sheets client for GKS
        client = get_google_sheets_client('ahly_gks')
        if not client:
            print("Google Sheets client not available for GKS")
            return False
        
        sheet_id = app.config['SHEET_IDS']['ahly_gks']
        worksheet_name = app.config['WORKSHEET_NAMES']['ahly_gks']
        
        # Open the spreadsheet
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get the worksheet, create if it doesn't exist
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)
            # Add headers for new worksheet
            headers = get_headers_for_type('ahly_gks')
            worksheet.append_row(headers)
        
        # Check if worksheet is empty, if so add headers
        all_values = worksheet.get_all_values()
        if not all_values:
            # Worksheet is empty, add headers
            headers = get_headers_for_type('ahly_gks')
            worksheet.append_row(headers)
        
        # Add a blank row before new data
        worksheet.append_row([''] * len(get_headers_for_type('ahly_gks')))
        
        # Extract GKS data
        match_id = data.get('match_id', '')
        gks_data = []
        
        # Find all GKS entries
        gks_count = 0
        while f'gks_{gks_count}_player_name' in data:
            player_name = data.get(f'gks_{gks_count}_player_name', '')
            eleven_backup = data.get(f'gks_{gks_count}_eleven_backup', '')
            submin = data.get(f'gks_{gks_count}_submin', '')
            team = data.get(f'gks_{gks_count}_team', '')
            goals_conceded = data.get(f'gks_{gks_count}_goals_conceded', '')
            goal_minute = data.get(f'gks_{gks_count}_goal_minute', '')
            
            # Only save if player name is provided
            if player_name:
                row_data = [match_id, player_name, eleven_backup, submin, team, goals_conceded, goal_minute]
                gks_data.append(row_data)
            
            gks_count += 1
        
        # Add GKS data to sheet
        for gk_row in gks_data:
            worksheet.append_row(gk_row)
        
        print(f"Successfully saved {len(gks_data)} GKS entries to GKDETAILS")
        return True
        
    except Exception as e:
        print(f"Error saving GKS data: {e}")
        return False

def save_howpenmissed_to_sheets(data):
    """Save HOWPENMISSED data to HOWPENMISSED sheet"""
    try:
        # Get Google Sheets client for HOWPENMISSED
        client = get_google_sheets_client('ahly_howpenmissed')
        if not client:
            print("Google Sheets client not available for HOWPENMISSED")
            return False
        
        sheet_id = app.config['SHEET_IDS']['ahly_howpenmissed']
        worksheet_name = app.config['WORKSHEET_NAMES']['ahly_howpenmissed']
        
        # Open the spreadsheet
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get the worksheet, create if it doesn't exist
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)
            # Add headers for new worksheet
            headers = get_headers_for_type('ahly_howpenmissed')
            worksheet.append_row(headers)
        
        # Check if worksheet is empty, if so add headers
        all_values = worksheet.get_all_values()
        if not all_values:
            # Worksheet is empty, add headers
            headers = get_headers_for_type('ahly_howpenmissed')
            worksheet.append_row(headers)
        
        # Add a blank row before new data
        worksheet.append_row([''] * len(get_headers_for_type('ahly_howpenmissed')))
        
        # Extract HOWPENMISSED data
        match_id = data.get('match_id', '')
        howpenmissed_data = []
        
        # Find all HOWPENMISSED entries
        howpenmissed_count = 0
        while f'howpenmissed_{howpenmissed_count}_player_name' in data:
            player_name = data.get(f'howpenmissed_{howpenmissed_count}_player_name', '')
            team = data.get(f'howpenmissed_{howpenmissed_count}_team', '')
            minute = data.get(f'howpenmissed_{howpenmissed_count}_minute', '')
            
            # Only save if player name is provided
            if player_name:
                row_data = [match_id, player_name, team, minute]
                howpenmissed_data.append(row_data)
            
            howpenmissed_count += 1
        
        # Add HOWPENMISSED data to sheet
        for row in howpenmissed_data:
            worksheet.append_row(row)
        
        print(f"Successfully saved {len(howpenmissed_data)} HOWPENMISSED entries to HOWPENMISSED")
        return True
        
    except Exception as e:
        print(f"Error saving HOWPENMISSED data: {e}")
        return False

def save_lineup_to_google_apps_script(data):
    """Save lineup data using Google Apps Script"""
    try:
        # Extract players data from form
        players_data = []
        player_count = 0
        
        while f'player_{player_count}_name' in data:
            player_data = {
                'name': data.get(f'player_{player_count}_name', ''),
                'minmat': data.get(f'player_{player_count}_minmat', ''),
                'status': data.get(f'player_{player_count}_status', ''),
                'playerout': data.get(f'player_{player_count}_playerout', ''),
                'minout': data.get(f'player_{player_count}_minout', ''),
                'mintotal': data.get(f'player_{player_count}_mintotal', '')
            }
            players_data.append(player_data)
            player_count += 1
        
        # Prepare data for Google Apps Script
        payload = {
            'target_sheet': data.get('target_sheet'),
            'match_date': data.get('match_date'),
            'match_id': data.get('match_id'),
            'players': players_data
        }
        
        # Get Google Apps Script URL from config
        script_url = app.config.get('GOOGLE_APPS_SCRIPT_URL')
        if not script_url:
            return False, "Google Apps Script URL not configured"
        
        # Send data to Google Apps Script
        print(f"Sending data to Google Apps Script: {script_url}")
        print(f"Payload: {json.dumps(payload, indent=2)}")
        
        response = requests.post(script_url, json=payload, timeout=30)
        
        print(f"Response status: {response.status_code}")
        print(f"Response content: {response.text}")
        
        if response.status_code == 200:
            try:
                result = response.json()
                if result.get('success'):
                    return True, result.get('message', 'تم الحفظ')
                else:
                    return False, result.get('message', 'Unknown error from Google Apps Script')
            except json.JSONDecodeError:
                return False, f"Invalid JSON response from Google Apps Script: {response.text}"
        else:
            return False, f"HTTP Error {response.status_code}: {response.text}"
            
    except Exception as e:
        return False, f"Error communicating with Google Apps Script: {str(e)}"

def save_ahly_pks_to_google_apps_script(data):
    """Save AHLY PKs data using Google Apps Script"""
    try:
        # Extract individual player fields for Google Apps Script
        ahly_count = 0
        while f'ahly_{ahly_count}_player_name' in data:
            data[f'ahly_{ahly_count}_player_name'] = data.get(f'ahly_{ahly_count}_player_name', '')
            data[f'ahly_{ahly_count}_eleven_backup'] = data.get(f'ahly_{ahly_count}_eleven_backup', '')
            data[f'ahly_{ahly_count}_howmiss'] = data.get(f'ahly_{ahly_count}_howmiss', '')
            ahly_count += 1
        
        opponent_count = 0
        while f'opponent_{opponent_count}_player_name' in data:
            data[f'opponent_{opponent_count}_player_name'] = data.get(f'opponent_{opponent_count}_player_name', '')
            data[f'opponent_{opponent_count}_eleven_backup'] = data.get(f'opponent_{opponent_count}_eleven_backup', '')
            data[f'opponent_{opponent_count}_howmiss'] = data.get(f'opponent_{opponent_count}_howmiss', '')
            opponent_count += 1
        
        # Prepare data for Google Apps Script - send all data including individual player fields
        payload = data.copy()  # Send all form data including individual player fields
        
        # Get Google Apps Script URL for PKs from config
        script_url = app.config.get('GOOGLE_APPS_SCRIPT_PKS_URL')
        if not script_url or script_url == 'https://script.google.com/macros/s/YOUR_PKS_SCRIPT_ID/exec':
            return False, "Google Apps Script PKs URL not configured. Please set GOOGLE_APPS_SCRIPT_PKS_URL in config.py"
        
        # Send data to Google Apps Script
        print(f"Sending PKs data to Google Apps Script: {script_url}")
        print(f"Payload: {json.dumps(payload, indent=2)}")
        
        response = requests.post(script_url, json=payload, timeout=30)
        
        print(f"Response status: {response.status_code}")
        print(f"Response content: {response.text}")
        
        if response.status_code == 200:
            try:
                result = response.json()
                if result.get('success'):
                    return True, result.get('message', 'تم الحفظ')
                else:
                    return False, result.get('message', 'Unknown error from Google Apps Script')
            except json.JSONDecodeError:
                return False, f"Invalid JSON response from Google Apps Script: {response.text}"
        else:
            return False, f"HTTP Error {response.status_code}: {response.text}"
            
    except Exception as e:
        return False, f"Error communicating with Google Apps Script: {str(e)}"

def get_sheet_data(data_type):
    """Get data from Google Sheet for export"""
    try:
        client = get_google_sheets_client(data_type)
        if not client:
            return None, "Google Sheets client not available"
        
        sheet_id = app.config['SHEET_IDS'][data_type]
        worksheet_name = app.config['WORKSHEET_NAMES'][data_type]
        
        # Open the spreadsheet
        spreadsheet = client.open_by_key(sheet_id)
        
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            return None, f"Worksheet '{worksheet_name}' not found"
        
        # Get all records
        records = worksheet.get_all_records()
        
        return records, "Data retrieved successfully"
        
    except Exception as e:
        return None, f"Error retrieving data: {str(e)}"

def create_excel_file(data_type, data):
    """Create Excel file from data"""
    try:
        # Define column order based on data type
        if data_type == 'ahly_lineup':
            columns = ['MATCH_ID', 'DATE', 'MINMAT', 'PLAYER', 'STATU', 'PLAYEROUT', 'MINOUT', 'MINTOTAL']
        elif data_type == 'ahly_match':
            columns = ['MATCH_ID', 'CHAMPION SYSTEM', 'DATE', 'CHAMPION', 'SEASON', 'AHLY MANAGER', 'OPPONENT MANAGER', 'REFREE', 'ROUND', 'H-A-N', 'STAD', 'AHLY TEAM', 'GF', 'GA', 'ET', 'PEN', 'OPPONENT TEAM', 'W-D-L', 'CLEAN SHEET', 'NOTE']
        elif data_type == 'egypt_match':
            columns = ['MATCH_ID', 'CHAMPION SYSTEM', 'DATE', 'MANAGER EGY', 'MANAGER OPPONENT', 'REFREE', 'CHAMPION', 'SEASON', 'ROUND', 'PLACE', 'H-A-N', 'Egypt TEAM', 'GF', 'GA', 'ET', 'PEN', 'OPPONENT TEAM', 'W-D-L', 'CLEAN SHEET', 'NOTE']
        elif data_type == 'egypt_lineup':
            columns = ['MATCH_ID', 'MINMAT', 'PLAYER NAME', 'STATU', 'PLAYER NAME OUT', 'MINOUT', 'MINTOTAL']
        else:
            columns = list(data[0].keys()) if data else []
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # If DataFrame is empty, create empty DataFrame with correct columns
        if df.empty:
            df = pd.DataFrame(columns=columns)
        else:
            # Reorder columns if they exist
            existing_columns = [col for col in columns if col in df.columns]
            if existing_columns:
                df = df[existing_columns]
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
        
        output.seek(0)
        return output, "Excel file created successfully"
        
    except Exception as e:
        return None, f"Error creating Excel file: {str(e)}"

def get_headers_for_type(data_type):
    """Get appropriate headers for each data type"""
    headers = {
        'ahly_match': ['MATCH_ID', 'CHAMPION SYSTEM', 'DATE', 'CHAMPION', 'SEASON', 'AHLY MANAGER', 'OPPONENT MANAGER', 'REFREE', 'ROUND', 'H-A-N', 'STAD', 'AHLY TEAM', 'GF', 'GA', 'ET', 'PEN', 'OPPONENT TEAM', 'W-D-L', 'CLEAN SHEET', 'NOTE'],
        'ahly_lineup': ['MATCH_ID', 'DATE', 'MINMAT', 'PLAYER', 'STATU', 'PLAYEROUT', 'MINOUT', 'MINTOTAL'],
        'ahly_goals_assists': ['MATCH_ID', 'PLAYER NAME', 'TEAM', 'GA', 'TYPE', 'MINUTE'],
        'ahly_gks': ['MATCH_ID', 'PLAYER NAME', '11/BAKEUP', 'SUBMIN', 'TEAM', 'GOALS CONCEDED', 'GOAL MINUTE'],
        'ahly_howpenmissed': ['MATCH_ID', 'PLAYER NAME', 'TEAM', 'MINUTE'],
        'egypt_match': ['MATCH_ID', 'CHAMPION SYSTEM', 'DATE', 'MANAGER EGY', 'MANAGER OPPONENT', 'REFREE', 'CHAMPION', 'SEASON', 'ROUND', 'PLACE', 'H-A-N', 'Egypt TEAM', 'GF', 'GA', 'ET', 'PEN', 'OPPONENT TEAM', 'W-D-L', 'CLEAN SHEET', 'NOTE'],
        'egypt_lineup': ['MATCH_ID', 'MINMAT', 'PLAYER NAME', 'STATU', 'PLAYER NAME OUT', 'MINOUT', 'MINTOTAL']
    }
    return headers.get(data_type, [])

def prepare_data_row(data_type, data):
    """Prepare data row based on type"""
    
    if data_type == 'ahly_match':
        return [
            data.get('match_id', ''),           # MATCH_ID
            data.get('champion_system', ''),    # CHAMPION SYSTEM
            data.get('date', ''),               # DATE
            data.get('champion', ''),           # CHAMPION
            data.get('season', ''),             # SEASON
            data.get('ahly_manager', ''),       # AHLY MANAGER
            data.get('opponent_manager', ''),   # OPPONENT MANAGER
            data.get('referee', ''),            # REFREE
            data.get('round', ''),              # ROUND
            data.get('h_a_n', ''),              # H-A-N
            data.get('stadium', ''),            # STAD
            data.get('ahly_team', ''),          # AHLY TEAM
            data.get('gf', ''),                 # GF
            data.get('ga', ''),                 # GA
            data.get('et', ''),                 # ET
            data.get('pen', ''),                # PEN
            data.get('opponent_team', ''),      # OPPONENT TEAM
            data.get('w_d_l', ''),              # W-D-L
            data.get('clean_sheet', ''),        # CLEAN SHEET
            data.get('notes', '')               # NOTE
        ]
    elif data_type == 'ahly_lineup':
        # Process multiple players for lineup
        players_data = []
        match_date = data.get('match_date', '')
        match_id = data.get('match_id', '')
        
        # Extract player data
        player_count = 0
        while f'player_{player_count}_name' in data:
            player_data = {
                'name': data.get(f'player_{player_count}_name', ''),
                'minmat': data.get(f'player_{player_count}_minmat', ''),
                'status': data.get(f'player_{player_count}_status', ''),
                'playerout': data.get(f'player_{player_count}_playerout', ''),
                'minout': data.get(f'player_{player_count}_minout', ''),
                'mintotal': data.get(f'player_{player_count}_mintotal', '')
            }
            players_data.append(player_data)
            player_count += 1
        
        # Return all players as separate rows
        return players_data, match_date, match_id
    elif data_type == 'egypt_lineup':
        # Process multiple players for Egypt lineup (same as Ahly lineup)
        players_data = []
        match_date = data.get('match_date', '')
        match_id = data.get('match_id', '')
        
        # Extract player data
        player_count = 0
        while f'player_{player_count}_name' in data:
            player_data = {
                'name': data.get(f'player_{player_count}_name', ''),
                'minmat': data.get(f'player_{player_count}_minmat', ''),
                'status': data.get(f'player_{player_count}_status', ''),
                'playerout': data.get(f'player_{player_count}_playerout', ''),
                'minout': data.get(f'player_{player_count}_minout', ''),
                'mintotal': data.get(f'player_{player_count}_mintotal', '')
            }
            players_data.append(player_data)
            player_count += 1
        
        # Return all players as separate rows
        return players_data, match_date, match_id
    elif data_type == 'egypt_match':
        return [
            data.get('match_id', ''),
            data.get('champion_system', ''),
            data.get('date', ''),
            data.get('manager_egy', ''),
            data.get('manager_opponent', ''),
            data.get('referee', ''),
            data.get('champion', ''),
            data.get('season', ''),
            data.get('round', ''),
            data.get('place', ''),
            data.get('h_a_n', ''),
            data.get('egypt', 'Egypt'),
            data.get('gf', ''),
            data.get('ga', ''),
            data.get('et', ''),
            data.get('pen', ''),
            data.get('opponent', ''),
            data.get('w_d_l', ''),
            data.get('cs', ''),
            data.get('note', '')
        ]
    
    return []

@app.route('/')
def index():
    return render_template('main_dashboard.html')

@app.route('/data-entry-login')
def data_entry_login():
    return render_template('data_entry_login.html')

@app.route('/al-ahly-stats')
def al_ahly_stats():
    return render_template('al_ahly_stats.html')

@app.route('/al-ahly-pks-stats')
def al_ahly_pks_stats():
    return render_template('al_ahly_pks_stats.html')

@app.route('/al-ahly-finals')
def al_ahly_finals():
    """Al Ahly Finals Statistics Page"""
    return render_template('al_ahly_finals.html')

@app.route('/ahly-match')
def ahly_match():
    return render_template('data_entry_ahly_match.html')

@app.route('/ahly-lineup')
def ahly_lineup():
    return render_template('data_entry_ahly_lineup.html')

@app.route('/ahly-pks')
def ahly_pks():
    return render_template('data_entry_ahly_pks.html')

@app.route('/ahly-vs-zamalek')
def ahly_vs_zamalek():
    return render_template('ahly_vs_zamalek.html')

@app.route('/egypt-teams')
def egypt_teams():
    """Egypt National Teams Statistics Page"""
    return render_template('egypt_teams.html')

@app.route('/youth-egypt-teams')
def youth_egypt_teams():
    """Youth Egypt National Teams Statistics Page"""
    return render_template('egypt_youth.html')

@app.route('/pks-egypt')
def pks_egypt():
    """PKS Egypt Statistics Page"""
    return render_template('egypt_pks.html')

@app.route('/egyptian-clubs')
def egyptian_clubs():
    """Egyptian Clubs Statistics Page"""
    return render_template('egyptian_clubs.html')

@app.route('/api/egyptian-clubs/config', methods=['GET'])
def api_egyptian_clubs_config():
    """Get Egyptian Clubs configuration (Apps Script URL)"""
    try:
        # Return the Apps Script URL
        apps_script_url = os.environ.get(
            'EGYPTIAN_CLUBS_APPS_SCRIPT_URL',
            'https://script.google.com/macros/s/AKfycby5gNCl_2Q-nNVktUAidJizj09WSthI7_tOx14hAw4KXGG8sHCICCRm9D1fTwg4HY0YaQ/exec'
        )
        
        return jsonify({
            'success': True,
            'appsScriptUrl': apps_script_url,
            'sheetId': '10UA-7awu0E_WBbxehNznng83MIUMVLCmpspvvkS1hTU'
        })
    except Exception as e:
        print(f"❌ Error getting Egyptian Clubs config: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/egypt-match')
def egypt_match():
    return render_template('data_entry_egypt_match.html')

@app.route('/egypt-lineup')
def egypt_lineup():
    return render_template('data_entry_egypt_lineup.html')

@app.route('/save_data', methods=['POST'])
def save_data():
    """Handle form submissions for all data types"""
    try:
        data_type = request.form.get('data_type')
        data = request.form.to_dict()
        
        # Remove data_type from the data dict
        data.pop('data_type', None)
        
        # Special handling for ahly_lineup with target sheet selection
        if data_type == 'ahly_lineup' and data.get('target_sheet'):
            success, message = save_lineup_to_google_apps_script(data)
        # Special handling for ahly_pks
        elif data_type == 'ahly_pks':
            success, message = save_ahly_pks_to_google_apps_script(data)
        else:
            success, message = save_to_sheets(data_type, data)
        
        if success:
            pass  # تم الحفظ
        else:
            pass  # فشل في الحفظ
            
    except Exception as e:
        pass  # فشل في الحفظ
    
    # Redirect to the appropriate tab based on data_type
    redirect_map = {
        'ahly_match': 'ahly_match',
        'ahly_lineup': 'ahly_lineup',
        'ahly_pks': 'ahly_pks',
        'egypt_match': 'egypt_match',
        'egypt_lineup': 'egypt_lineup'
    }
    
    return redirect(url_for(redirect_map.get(data_type, 'ahly_match')))

@app.route('/api/save_data', methods=['POST'])
def api_save_data():
    """API endpoint for saving data"""
    try:
        data = request.get_json()
        data_type = data.get('data_type')
        
        if not data_type:
            return jsonify({'success': False, 'message': 'Data type is required'}), 400
        
        # Special handling for AHLY PKs
        if data_type == 'ahly_pks':
            success, message = save_ahly_pks_to_google_apps_script(data)
        else:
            success, message = save_to_sheets(data_type, data)
        
        return jsonify({'success': success, 'message': message})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/create_excel', methods=['POST'])
def create_excel_from_data():
    """Create Excel file from provided data"""
    try:
        data = request.get_json()
        data_type = data.get('data_type')
        excel_data = data.get('data', [])
        
        if not data_type:
            return jsonify({'error': 'Data type is required'}), 400
        
        # Create Excel file
        excel_file, message = create_excel_file(data_type, excel_data)
        
        if excel_file is None:
            return jsonify({'error': message}), 500
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{data_type}_{timestamp}.xlsx'
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/<data_type>')
def export_data(data_type):
    """Export data to Excel file"""
    try:
        # For AHLY MATCH, create empty template (form data will be handled by form submission)
        if data_type == 'ahly_match':
            data = create_empty_template(data_type)
        else:
            # Try to get data from Google Sheets first for other data types
            data, message = get_sheet_data(data_type)
        
        if data is None:
            # If Google Sheets is not available, create empty template
            # Google Sheets not available. Creating empty template.
            data = []
        
        if not data:
            # Create empty template with headers
            data = create_empty_template(data_type)
        
        # Create Excel file
        excel_file, message = create_excel_file(data_type, data)
        
        if excel_file is None:
            # فشل في الحفظ
            return redirect(url_for('ahly_match'))
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{data_type}_{timestamp}.xlsx'
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # فشل في الحفظ
        return redirect(url_for('ahly_match'))

@app.route('/export_form/<data_type>', methods=['POST'])
def export_form_data(data_type):
    """Export form data to Excel file"""
    try:
        # Get form data
        form_data = request.form.to_dict()
        
        # Convert form data to the format expected by create_excel_file
        if data_type == 'ahly_match':
            # Map form field names to column names for Match sheet
            field_mapping = {
                'match_id': 'MATCH_ID',
                'champion_system': 'CHAMPION SYSTEM',
                'date': 'DATE',
                'champion': 'CHAMPION',
                'season': 'SEASON',
                'ahly_manager': 'AHLY MANAGER',
                'opponent_manager': 'OPPONENT MANAGER',
                'referee': 'REFREE',
                'round': 'ROUND',
                'h_a_n': 'H-A-N',
                'stadium': 'STAD',
                'ahly_team': 'AHLY TEAM',
                'gf': 'GF',
                'ga': 'GA',
                'et': 'ET',
                'pen': 'PEN',
                'opponent_team': 'OPPONENT TEAM',
                'w_d_l': 'W-D-L',
                'clean_sheet': 'CLEAN SHEET',
                'notes': 'NOTE'
            }
            
            # Convert form data to proper column names for Match sheet
            converted_data = {}
            for form_field, value in form_data.items():
                if form_field in field_mapping:
                    converted_data[field_mapping[form_field]] = value
            
            # Create Excel file with two sheets
            excel_file = create_ahly_match_excel_with_sheets(converted_data, form_data)
            
            if excel_file is None:
                pass  # Save failed
                return redirect(url_for('ahly_match'))
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{data_type}_{timestamp}.xlsx'
            
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        elif data_type == 'egypt_match':
            # Map form field names to column names for Egypt Match
            field_mapping = {
                'champion_system': 'CHAMPION SYSTEM',
                'date': 'DATE',
                'manager_egy': 'MANAGER EGY',
                'manager_opponent': 'MANAGER OPPONENT',
                'referee': 'REFREE',
                'champion': 'CHAMPION',
                'season': 'SEASON',
                'round': 'ROUND',
                'place': 'PLACE',
                'h_a_n': 'H-A-N',
                'egypt': 'Egypt',
                'gf': 'GF',
                'ga': 'GA',
                'et': 'ET',
                'pen': 'PEN',
                'opponent': 'OPPONENT',
                'w_d_l': 'W-D-L',
                'cs': 'CS',
                'note': 'NOTE'
            }
            
            # Convert form data to proper column names
            converted_data = {}
            for form_field, value in form_data.items():
                if form_field in field_mapping:
                    converted_data[field_mapping[form_field]] = value
            
            # Check if we have data
            if not converted_data:
                pass  # Save failed
                return redirect(url_for('egypt_match'))
            
            # Create Excel file
            excel_file, message = create_excel_file(data_type, [converted_data])
            
            if excel_file is None:
                pass  # Save failed
                return redirect(url_for('egypt_match'))
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{data_type}_{timestamp}.xlsx'
            
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # For other data types, create empty template
            data = create_empty_template(data_type)
            
            # Create Excel file
            excel_file, message = create_excel_file(data_type, data)
            
            if excel_file is None:
                pass  # Save failed
                return redirect(url_for('ahly_match'))
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{data_type}_{timestamp}.xlsx'
            
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    except Exception as e:
        # Save failed
        return redirect(url_for('ahly_match'))

def create_ahly_match_excel_with_sheets(match_data, form_data):
    """Create Excel file with two sheets: Match and Goals & Assists"""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Match sheet
        match_ws = wb.create_sheet("Match")
        
        # Add headers for Match sheet
        match_headers = ['MATCH_ID', 'CHAMPION SYSTEM', 'DATE', 'CHAMPION', 'SEASON', 'AHLY MANAGER', 'OPPONENT MANAGER', 'REFREE', 'ROUND', 'H-A-N', 'STAD', 'AHLY TEAM', 'GF', 'GA', 'ET', 'PEN', 'OPPONENT TEAM', 'W-D-L', 'CLEAN SHEET', 'NOTE']
        match_ws.append(match_headers)
        
        # Add match data
        match_row = []
        for header in match_headers:
            match_row.append(match_data.get(header, ''))
        match_ws.append(match_row)
        
        # Create Goals & Assists sheet
        goals_ws = wb.create_sheet("Goals & Assists")
        
        # Add headers for Goals & Assists sheet
        goals_headers = ['MATCH_ID', 'PLAYER NAME', 'TEAM', 'GA', 'TYPE', 'MINUTE']
        goals_ws.append(goals_headers)
        
        # Extract Goals & Assists data
        goals_data = []
        match_id = match_data.get('MATCH_ID', '')
        
        # Find all goals & assists entries
        goals_count = 0
        while f'goals_assists_{goals_count}_player_name' in form_data:
            player_name = form_data.get(f'goals_assists_{goals_count}_player_name', '')
            team = form_data.get(f'goals_assists_{goals_count}_team', '')
            ga = form_data.get(f'goals_assists_{goals_count}_ga', '')
            goal_type = form_data.get(f'goals_assists_{goals_count}_type', '')
            minute = form_data.get(f'goals_assists_{goals_count}_minute', '')
            
            # Only add if player name is provided
            if player_name:
                goals_data.append([match_id, player_name, team, ga, goal_type, minute])
            
            goals_count += 1
        
        # Add goals data to sheet
        for goal_row in goals_data:
            goals_ws.append(goal_row)
        
        # Create GKS sheet
        gks_ws = wb.create_sheet("GKS")
        
        # Add headers for GKS sheet
        gks_headers = ['MATCH_ID', 'PLAYER NAME', '11/BAKEUP', 'SUBMIN', 'TEAM', 'GOALS CONCEDED', 'GOAL MINUTE']
        gks_ws.append(gks_headers)
        
        # Extract GKS data
        gks_data = []
        
        # Find all GKS entries
        gks_count = 0
        while f'gks_{gks_count}_player_name' in form_data:
            player_name = form_data.get(f'gks_{gks_count}_player_name', '')
            eleven_backup = form_data.get(f'gks_{gks_count}_eleven_backup', '')
            submin = form_data.get(f'gks_{gks_count}_submin', '')
            team = form_data.get(f'gks_{gks_count}_team', '')
            goals_conceded = form_data.get(f'gks_{gks_count}_goals_conceded', '')
            goal_minute = form_data.get(f'gks_{gks_count}_goal_minute', '')
            
            # Only add if player name is provided
            if player_name:
                gks_data.append([match_id, player_name, eleven_backup, submin, team, goals_conceded, goal_minute])
            
            gks_count += 1
        
        # Add GKS data to sheet
        for gk_row in gks_data:
            gks_ws.append(gk_row)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        print(f"Error creating Excel with sheets: {e}")
        return None

def create_empty_template(data_type):
    """Create empty template with headers for export"""
    if data_type == 'ahly_lineup':
        return [{
            'DATE': '',
            'MATCH_ID': '',
            'MINMAT': '',
            'PLAYER': '',
            'STATU': '',
            'PLAYEROUT': '',
            'MINOUT': '',
            'MINTOTAL': ''
        }]
    elif data_type == 'ahly_match':
        return [{
            'Match ID': '',
            'Date': '',
            'Champion System': '',
            'H-A-N': '',
            'Stadium': '',
            'Champion': '',
            'Season': '',
            'Round': '',
            'Ahly Manager': '',
            'Opponent Manager': '',
            'Referee': '',
            'Ahly Team': '',
            'Goals For': '',
            'Goals Against': '',
            'Opponent Team': '',
            'Extra Time': '',
            'Penalties': '',
            'Result': '',
            'Clean Sheet': '',
            'Notes': ''
        }]
    elif data_type == 'egypt_match':
        return [{
            'Date': '',
            'Competition': '',
            'Opponent': '',
            'Egypt Score': '',
            'Opponent Score': '',
            'Venue': '',
            'Notes': ''
        }]
    elif data_type == 'egypt_lineup':
        return [{
            'MATCH_ID': '',
            'MINMAT': '',
            'PLAYER NAME': '',
            'STATU': '',
            'PLAYER NAME OUT': '',
            'MINOUT': '',
            'MINTOTAL': ''
        }]
    else:
        return []

@app.route('/api/players')
def get_players():
    """Get player names from PLAYERDATABASE sheet (Ahly Sheet)"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('ahly_players_list', ttl_hours=6)
        if cached_data:
            return jsonify({'players': cached_data})
        
        # Use Ahly credentials (sheet ID: 1zeSlEN7VS2S6KPZH7_uvQeeY3Iu5INUyi12V0_Wi9G4)
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for Ahly (which contains PLAYERDATABASE)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get PLAYERDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'PLAYERDATABASE worksheet not found'}), 404
        
        # Get all values from the sheet
        all_values = worksheet.get_all_values()
        if not all_values:
            return jsonify({'players': []})
        
        # Get the first row as headers
        headers = all_values[0]
        
        # Find the PLAYER NAME column index
        try:
            player_name_index = headers.index('PLAYER NAME')
        except ValueError:
            return jsonify({'error': 'PLAYER NAME column not found'}), 404
        
        # Extract player names from all rows (excluding header)
        players = []
        for row in all_values[1:]:
            if len(row) > player_name_index and row[player_name_index].strip():
                players.append(row[player_name_index].strip())
        
        # Remove duplicates and sort
        players = sorted(list(set(players)))
        
        # Cache the result
        cache.set('ahly_players_list', players)
        
        return jsonify({'players': players})
        
    except Exception as e:
        print(f"Error getting players: {e}")
        return jsonify({'error': 'Failed to get players'}), 500

@app.route('/api/egypt-players')
def get_egypt_players():
    """Get player names from PLAYERDATABASE sheet (Egypt Sheet)"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('egypt_players_list', ttl_hours=6)
        if cached_data:
            return jsonify({'players': cached_data})
        
        # Use Egypt Teams credentials (sheet ID: 10PbAfoH9eqr4F82EBtO281RO42DgRzUzRv-dtELRDn8)
        client = get_google_sheets_client('egypt_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for Egypt (which contains PLAYERDATABASE)
        sheet_id = app.config['SHEET_IDS']['egypt_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get PLAYERDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'PLAYERDATABASE worksheet not found'}), 404
        
        # Get all values from the sheet
        all_values = worksheet.get_all_values()
        if not all_values:
            return jsonify({'players': []})
        
        # Get the first row as headers
        headers = all_values[0]
        
        # Find the PLAYER NAME column index
        try:
            player_name_index = headers.index('PLAYER NAME')
        except ValueError:
            return jsonify({'error': 'PLAYER NAME column not found'}), 404
        
        # Extract player names from all rows (excluding header)
        players = []
        for row in all_values[1:]:
            if len(row) > player_name_index and row[player_name_index].strip():
                players.append(row[player_name_index].strip())
        
        # Remove duplicates and sort
        players = sorted(list(set(players)))
        
        # Cache the result
        cache.set('egypt_players_list', players)
        
        return jsonify({'players': players})
        
    except Exception as e:
        print(f"Error getting Egypt players: {e}")
        return jsonify({'error': 'Failed to get Egypt players'}), 500

@app.route('/api/teams')
def get_teams():
    """Get team names from TEAMDATABASE sheet"""
    try:
        print("🎯 API CALL: /api/teams")
        
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('teams_list', ttl_hours=6)
        if cached_data:
            return jsonify({'teams': cached_data})
        
        # Use the same credentials as ahly_match
        client = get_google_sheets_client('ahly_match')
        if not client:
            print("❌ Google Sheets client not available")
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for ahly_match (which contains TEAMDATABASE)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get TEAMDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('TEAMDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'TEAMDATABASE worksheet not found'}), 404
        
        # Get all values from the sheet
        all_values = worksheet.get_all_values()
        if not all_values:
            return jsonify({'teams': []})
        
        # Get the first row as headers
        headers = all_values[0]
        
        # Find the TEAM NAME column index
        try:
            team_name_index = headers.index('TEAM NAME')
        except ValueError:
            return jsonify({'error': 'TEAM NAME column not found'}), 404
        
        # Extract team names from all rows (excluding header)
        teams = []
        for row in all_values[1:]:
            if len(row) > team_name_index and row[team_name_index].strip():
                teams.append(row[team_name_index].strip())
        
        # Remove duplicates and sort
        teams = sorted(list(set(teams)))
        
        # Cache the result
        cache.set('teams_list', teams)
        
        print(f"✅ Loaded {len(teams)} teams from TEAMDATABASE")
        
        return jsonify({'teams': teams})
        
    except Exception as e:
        print(f"Error getting teams: {e}")
        return jsonify({'error': 'Failed to get teams'}), 500

@app.route('/api/goal-types')
def get_goal_types():
    """Get goal types from PLAYERDETAILS sheet"""
    try:
        
        # Use the same credentials as ahly_match
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for ahly_match (which contains PLAYERDETAILS)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get PLAYERDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'PLAYERDETAILS worksheet not found'}), 404
        
        # Get all values from the sheet
        all_values = worksheet.get_all_values()
        if not all_values:
            return jsonify({'types': []})
        
        # Get the first row as headers
        headers = all_values[0]
        
        # Find the TYPE column index
        try:
            type_index = headers.index('TYPE')
        except ValueError:
            return jsonify({'error': 'TYPE column not found'}), 404
        
        # Extract types from all rows (excluding header)
        types = []
        for row in all_values[1:]:
            if len(row) > type_index and row[type_index].strip():
                types.append(row[type_index].strip())
        
        # Remove duplicates and sort
        types = sorted(list(set(types)))
        
        
        return jsonify({'types': types})
        
    except Exception as e:
        print(f"Error getting goal types: {e}")
        return jsonify({'error': 'Failed to get goal types'}), 500

@app.route('/api/stadiums')
def get_stadiums():
    """Get stadium names from STADDATABASE sheet"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('stadiums_list', ttl_hours=6)
        if cached_data:
            return jsonify({'stadiums': cached_data})
        
        # Use the same credentials as ahly_match
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for ahly_match (which contains STADDATABASE)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get STADDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('STADDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'STADDATABASE worksheet not found'}), 404
        
        # Get all values from the sheet
        all_values = worksheet.get_all_values()
        if not all_values:
            return jsonify({'stadiums': []})
        
        # Get the first row as headers
        headers = all_values[0]
        
        # Find the STADIUM NAME column index
        try:
            stadium_name_index = headers.index('STADIUM NAME')
        except ValueError:
            return jsonify({'error': 'STADIUM NAME column not found'}), 404
        
        # Extract stadium names from all rows (excluding header)
        stadiums = []
        for row in all_values[1:]:
            if len(row) > stadium_name_index and row[stadium_name_index].strip():
                stadiums.append(row[stadium_name_index].strip())
        
        # Remove duplicates and sort
        stadiums = sorted(list(set(stadiums)))
        
        # Cache the result
        cache.set('stadiums_list', stadiums)
        
        return jsonify({'stadiums': stadiums})
        
    except Exception as e:
        print(f"Error getting stadiums: {e}")
        return jsonify({'error': 'Failed to get stadiums'}), 500

@app.route('/api/champions')
def get_champions():
    """Get champion names from MATCHDETAILS sheet"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('champions_list', ttl_hours=6)
        if cached_data:
            return jsonify({'champions': cached_data})
        
        # Use the same credentials as ahly_match
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for ahly_match (which contains MATCHDETAILS)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get MATCHDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('MATCHDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get all records
        records = worksheet.get_all_records()
        
        # Extract unique champions
        champions = []
        for record in records:
            champion = record.get('CHAMPION', '').strip()
            if champion and champion not in champions:
                champions.append(champion)
        
        # Remove duplicates and sort
        champions = sorted(list(set(champions)))
        
        # Cache the result
        cache.set('champions_list', champions)
        
        return jsonify({'champions': champions})
        
    except Exception as e:
        print(f"Error getting champions: {e}")
        return jsonify({'error': f'Failed to get champions: {str(e)}'}), 500

@app.route('/api/managers')
def get_managers():
    """Get manager names from MANAGERDATABASE sheet"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('managers_list', ttl_hours=6)
        if cached_data:
            return jsonify({'managers': cached_data})
        
        # Use the same credentials as ahly_match
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for ahly_match (which contains MANAGERDATABASE)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get MANAGERDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('MANAGERDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'MANAGERDATABASE worksheet not found'}), 404
        
        # Get all records
        records = worksheet.get_all_records()
        
        # Extract unique managers
        managers = []
        for record in records:
            manager = record.get('MANAGER NAME', '').strip()
            if manager and manager not in managers:
                managers.append(manager)
        
        # Remove duplicates and sort
        managers = sorted(list(set(managers)))
        
        # Cache the result
        cache.set('managers_list', managers)
        
        return jsonify({'managers': managers})
        
    except Exception as e:
        print(f"Error getting managers: {e}")
        return jsonify({'error': f'Failed to get managers: {str(e)}'}), 500

@app.route('/api/referees')
def get_referees():
    """Get referee names from RefereeDATABASE sheet"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('referees_list', ttl_hours=6)
        if cached_data:
            return jsonify({'referees': cached_data})
        
        # Use the same credentials as ahly_match
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for ahly_match (which contains RefereeDATABASE)
        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Try to get RefereeDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('RefereeDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'RefereeDATABASE worksheet not found'}), 404
        
        # Get all values from the sheet
        all_values = worksheet.get_all_values()
        if not all_values:
            return jsonify({'referees': []})
        
        # Get the first row as headers
        headers = all_values[0]
        
        # Find the REFEREE NAME column index
        try:
            referee_name_index = headers.index('REFEREE NAME')
        except ValueError:
            return jsonify({'error': 'REFEREE NAME column not found'}), 404
        
        # Extract referee names from all rows (excluding header)
        referees = []
        for row in all_values[1:]:
            if len(row) > referee_name_index and row[referee_name_index].strip():
                referees.append(row[referee_name_index].strip())
        
        # Remove duplicates and sort
        referees = sorted(list(set(referees)))
        
        # Cache the result
        cache.set('referees_list', referees)
        
        return jsonify({'referees': referees})
        
    except Exception as e:
        print(f"Error getting referees: {e}")
        return jsonify({'error': 'Failed to get referees'}), 500


@app.route('/api/players-data')
def api_players_data():
    """API endpoint to get players data from PLAYERDATABASE sheet"""
    try:
        # Backend Excel disabled for Ahly Stats page; return empty dataset
        return jsonify({'players': [], 'total_players': 0})
        
    except Exception as e:
        print(f"Error fetching players data: {e}")
        return jsonify({'error': 'Failed to fetch players data'}), 500

@app.route('/api/player-all-stats/<player_name>')
def api_player_all_stats(player_name):
    """Return all player statistics in one API call"""
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading ALL stats for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Get all data in one call
        result = {
            'success': True,
            'player_name': player_name,
            'team_filter': team_filter,
            'overview_stats': {},
            'matches': [],
            'championships': [],
            'seasons': [],
            'vs_teams': [],
            'vs_goalkeepers': []
        }
        
        # Load overview stats from Google Sheets
        try:
            client = get_google_sheets_client('ahly_match')
            if not client:
                result['overview_stats'] = {'error': 'Google Sheets client not available'}
            else:
                spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
                
                # Get player details for stats calculation
                player_sheet = spreadsheet.worksheet('PLAYERDETAILS')
                player_records = player_sheet.get_all_records()
                
                # Get lineup details for minutes
                lineup_sheet = spreadsheet.worksheet('LINEUPDETAILS')
                lineup_records = lineup_sheet.get_all_records()
                
                # Calculate stats
                total_matches = 0
                total_goals = 0
                total_assists = 0
                total_minutes = 0
                
                # Count goals and assists
                for record in player_records:
                    if record.get('PLAYER NAME', '').strip() == player_name:
                        if team_filter and record.get('TEAM', '').strip() != team_filter:
                            continue
                        if record.get('GA', '').strip().upper() == 'GOAL':
                            total_goals += 1
                        elif record.get('GA', '').strip().upper() == 'ASSIST':
                            total_assists += 1
                
                # Count matches and minutes
                match_ids = set()
                for record in lineup_records:
                    if record.get('PLAYER NAME', '').strip() == player_name:
                        if team_filter and record.get('TEAM', '').strip() != team_filter:
                            continue
                        match_id = record.get('MATCH_ID', '').strip()
                        if match_id:
                            match_ids.add(match_id)
                        minutes = record.get('MINTOTAL', 0)
                        if minutes:
                            try:
                                total_minutes += int(minutes)
                            except (ValueError, TypeError):
                                pass
                
                total_matches = len(match_ids)
                
                result['overview_stats'] = {
                    'total_matches': total_matches,
                    'total_goals': total_goals,
                    'total_assists': total_assists,
                    'total_minutes': total_minutes
                }
        except Exception as e:
            print(f"Error loading overview stats: {e}")
            result['overview_stats'] = {'error': str(e)}
        
        # Load matches from Google Sheets
        try:
            client = get_google_sheets_client('ahly_match')
            if not client:
                result['matches'] = []
            else:
                spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
                
                # Get player details
                player_sheet = spreadsheet.worksheet('PLAYERDETAILS')
                player_records = player_sheet.get_all_records()
                
                # Get match details
                match_sheet = spreadsheet.worksheet('MATCHDETAILS')
                match_records = match_sheet.get_all_records()
                
                # Get lineup details
                lineup_sheet = spreadsheet.worksheet('LINEUPDETAILS')
                lineup_records = lineup_sheet.get_all_records()
                
                # Group by match
                match_data = {}
                for record in player_records:
                    if record.get('PLAYER NAME', '').strip() == player_name:
                        if team_filter and record.get('TEAM', '').strip() != team_filter:
                            continue
                        match_id = record.get('MATCH_ID', '').strip()
                        if match_id:
                            if match_id not in match_data:
                                match_data[match_id] = {'goals': 0, 'assists': 0, 'minutes': 0}
                            
                            if record.get('GA', '').strip().upper() == 'GOAL':
                                match_data[match_id]['goals'] += 1
                            elif record.get('GA', '').strip().upper() == 'ASSIST':
                                match_data[match_id]['assists'] += 1
                
                # Add minutes from lineup
                for record in lineup_records:
                    if record.get('PLAYER NAME', '').strip() == player_name:
                        if team_filter and record.get('TEAM', '').strip() != team_filter:
                            continue
                        match_id = record.get('MATCH_ID', '').strip()
                        if match_id in match_data:
                            minutes = record.get('MINTOTAL', 0)
                            if minutes:
                                try:
                                    match_data[match_id]['minutes'] = int(minutes)
                                except (ValueError, TypeError):
                                    pass
                
                # Build matches list
                matches = []
                for match_id, stats in match_data.items():
                    # Find match details
                    match_details = next((m for m in match_records if m.get('MATCH_ID', '') == match_id), {})
                    
                    matches.append({
                        'match_id': match_id,
                        'date': match_details.get('DATE', ''),
                        'season': match_details.get('SEASON', ''),
                        'ahly_manager': match_details.get('AHLY MANAGER', ''),
                        'opponent_team': match_details.get('OPPONENT TEAM', ''),
                        'goals': stats['goals'],
                        'assists': stats['assists'],
                        'minutes': stats['minutes']
                    })
                
                result['matches'] = matches
        except Exception as e:
            print(f"Error loading matches: {e}")
            result['matches'] = []
        
        # Load championships (placeholder for now)
        result['championships'] = []
        
        # Load seasons (placeholder for now)
        result['seasons'] = []
        
        # Load vs teams (placeholder for now)
        result['vs_teams'] = []
        
        # Load vs goalkeepers (placeholder for now)
        result['vs_goalkeepers'] = []
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error loading all player stats: {e}")
        return jsonify({'error': f'Failed to load player stats: {str(e)}'}), 500

@app.route('/api/player-overview-stats/<player_name>')
def api_player_overview_stats(player_name):
    """Return player overview statistics for a specific player"""
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading overview stats for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Backend data sources disabled for this page (Excel handled on frontend)
        player_records = []
        lineup_records = []
        
        # Calculate statistics
        total_goals = 0
        total_assists = 0
        total_matches = 0
        total_minutes = 0
        
        # Detailed statistics
        brace_goals = 0
        brace_assists = 0
        hat_trick_goals = 0
        hat_trick_assists = 0
        three_plus_goals = 0
        three_plus_assists = 0
        pen_goal = 0
        pen_ast_goal = 0
        pen_missed = 0
        pen_ast_miss = 0
        pen_make_g = 0
        pen_make_m = 0
        
        # Count goals and assists from PLAYERDETAILS
        for record in player_records:
            record_player_name = record.get('PLAYER NAME', '').strip()
            if record_player_name == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                ga_type = record.get('GA', '').strip().upper()
                
                # عد الأهداف والمساعدات الأساسية فقط (الكلمة كاملة)
                if ga_type == 'GOAL':
                    total_goals += 1
                elif ga_type == 'ASSIST':
                    total_assists += 1
                
                # عد ضربات الجزاء منفصلة (مش بتأثر على العدد الأساسي)
                elif ga_type == 'PENGOAL':
                    pen_goal += 1
                elif ga_type == 'PENASSISTGOAL':
                    pen_ast_goal += 1
                elif ga_type == 'PENMISSED':
                    pen_missed += 1
                elif ga_type == 'PENASSISTMISSED':
                    pen_ast_miss += 1
                elif ga_type == 'PENMAKEGOAL':
                    pen_make_g += 1
                elif ga_type == 'PENMAKEMISSED':
                    pen_make_m += 1
        
        # Count matches and minutes from LINEUPDETAILS
        match_ids = set()
        for record in lineup_records:
            record_player_name = record.get('PLAYER NAME', '').strip()
            if record_player_name == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                match_id = record.get('MATCH_ID', '').strip()
                if match_id:
                    match_ids.add(match_id)
                minutes = record.get('MINTOTAL', 0)
                if minutes:
                    try:
                        total_minutes += int(minutes)
                    except (ValueError, TypeError):
                        pass
        
        total_matches = len(match_ids)
        
        # Calculate BRACE and HAT TRICK statistics
        # Group goals and assists by match
        match_goals = {}
        match_assists = {}
        
        for record in player_records:
            if record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                match_id = record.get('MATCH_ID', '').strip()
                if not match_id:
                    continue
                    
                ga_type = record.get('GA', '').strip().upper()
                if ga_type in ['GOAL', 'PENGOAL', 'PENMAKEGOAL']:
                    if match_id not in match_goals:
                        match_goals[match_id] = 0
                    match_goals[match_id] += 1
                elif ga_type in ['ASSIST', 'PENASSISTGOAL', 'PENMAKEMISSED']:
                    if match_id not in match_assists:
                        match_assists[match_id] = 0
                    match_assists[match_id] += 1
        
        # Count BRACE and HAT TRICK
        for match_id, goals in match_goals.items():
            if goals == 2:
                brace_goals += 1
            elif goals == 3:
                hat_trick_goals += 1
            elif goals >= 3:
                three_plus_goals += 1
                
        for match_id, assists in match_assists.items():
            if assists == 2:
                brace_assists += 1
            elif assists == 3:
                hat_trick_assists += 1
            elif assists >= 3:
                three_plus_assists += 1
        
        stats = {
            'total_matches': total_matches,
            'total_goals': total_goals,
            'total_assists': total_assists,
            'total_minutes': total_minutes,
            'brace_goals': brace_goals,
            'brace_assists': brace_assists,
            'hat_trick_goals': hat_trick_goals,
            'hat_trick_assists': hat_trick_assists,
            'three_plus_goals': three_plus_goals,
            'three_plus_assists': three_plus_assists,
            'pen_goal': pen_goal,
            'pen_ast_goal': pen_ast_goal,
            'pen_missed': pen_missed,
            'pen_ast_miss': pen_ast_miss,
            'pen_make_g': pen_make_g,
            'pen_make_m': pen_make_m
        }
        
        result = {
            'success': True,
            'stats': stats
        }
        
        print(f"Player stats calculated: {stats}")
        return jsonify(result)
        
    except Exception as e:
        print(f"Error loading player overview stats: {e}")
        return jsonify({'error': f'Failed to load player stats: {str(e)}'}), 500

@app.route('/api/player-matches/<player_name>')
def api_player_matches(player_name):
    """Return player matches for a specific player"""
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading matches for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Check if Excel file exists first
        if excel_service.file_exists():
            # Use Excel data
            player_df = excel_service.get_dataframe('PLAYERDETAILS')
            match_df = excel_service.get_dataframe('MATCHDETAILS')
            lineup_df = excel_service.get_dataframe('LINEUPDETAILS')
            
            if player_df.empty:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found or empty'}), 404
            if match_df.empty:
                return jsonify({'error': 'MATCHDETAILS worksheet not found or empty'}), 404
            if lineup_df.empty:
                return jsonify({'error': 'LINEUPDETAILS worksheet not found or empty'}), 404
            
            # Convert to records format for compatibility
            player_records = player_df.to_dict('records')
            match_records = match_df.to_dict('records')
            lineup_records = lineup_df.to_dict('records')
            
            print(f"🔍 MATCHDETAILS columns: {list(match_records[0].keys()) if match_records else 'No records'}")
        
        else:
            # Fallback to Google Sheets
            client = get_google_sheets_client('ahly_match')
            if not client:
                return jsonify({'error': 'No Excel file uploaded and Google Sheets client not available'}), 500

            spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
            
            # Get player details for goals and assists
            try:
                player_sheet = spreadsheet.worksheet('PLAYERDETAILS')
                player_records = player_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found'}), 404
        
            # Get match details for match information
            try:
                match_sheet = spreadsheet.worksheet('MATCHDETAILS')
                match_records = match_sheet.get_all_records()
                print(f"🔍 MATCHDETAILS columns: {list(match_records[0].keys()) if match_records else 'No records'}")
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get lineup details for minutes
        try:
            lineup_sheet = spreadsheet.worksheet('LINEUPDETAILS')
            lineup_records = lineup_sheet.get_all_records()
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'LINEUPDETAILS worksheet not found'}), 404
        
        # Find matches where player scored goals or made assists
        player_ga_records = []
        print(f"🔍 Searching for player '{player_name}' in {len(player_records)} PLAYERDETAILS records")
        print(f"🔍 First few player names: {[r.get('PLAYER NAME', '').strip() for r in player_records[:5]]}")
        
        for record in player_records:
            if record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                ga_type = record.get('GA', '').strip().upper()
                print(f"🔍 Found record for '{player_name}': GA={ga_type}")
                if ga_type in ['GOAL', 'ASSIST']:
                    player_ga_records.append(record)
        
        print(f"📊 Found {len(player_ga_records)} GA records for player '{player_name}'")
        
        # Build matches data
        matches = []
        print(f"🏗️ Building matches data from {len(player_ga_records)} GA records")
        
        for ga_record in player_ga_records:
            match_id = ga_record.get('MATCH_ID', '')
            if not match_id:
                print(f"⚠️ Skipping record with no MATCH_ID")
                continue
            
            print(f"🔍 Looking for match_id: '{match_id}' in {len(match_records)} match records")
                
            # Find match details
            match_info = None
            for match_record in match_records:
                if match_record.get('MATCH_ID', '').strip() == match_id:
                    match_info = match_record
                    print(f"✅ Found match info for {match_id}")
                    break
            
            if not match_info:
                print(f"❌ No match info found for match_id: {match_id}")
                continue
            
            # Find lineup details for minutes
            minutes = 0
            for lineup_record in lineup_records:
                if (lineup_record.get('MATCH_ID', '').strip() == match_id and 
                    lineup_record.get('PLAYER NAME', '').strip() == player_name):
                    try:
                        minutes = int(lineup_record.get('MINTOTAL', 0) or 0)
                    except (ValueError, TypeError):
                        minutes = 0
                    break
            
            if match_info:
                match_data = {
                    'match_id': match_id,
                    'player_name': player_name,
                    'team': ga_record.get('TEAM', ''),
                    'opponent': match_info.get('OPPONENT TEAM', ''),
                    'date': match_info.get('DATE', ''),
                    'season': match_info.get('SEASON', ''),
                    'manager': match_info.get('AHLY MANAGER', ''),
                    'ga': ga_record.get('GA', ''),
                    'minutes': minutes
            }
            matches.append(match_data)
        
        # Sort matches by date (newest first)
        matches.sort(key=lambda x: x.get('date', ''), reverse=True)
        
        result = {
            'success': True,
            'player_name': player_name,
            'matches': matches,
            'total_matches': len(matches)
        }
        
        print(f"📊 Final result: {len(matches)} matches for '{player_name}'")
        if matches:
            print(f"📊 First match: {matches[0]}")
        
        return jsonify(result)
    except Exception as e:
        print(f"Error loading player matches: {e}")
        return jsonify({'error': f'Failed to load player matches: {str(e)}'}), 500

@app.route('/api/player-championships/<player_name>')
def api_player_championships(player_name):
    """API endpoint to get player championships from Google Sheets"""
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading championships for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Check if Excel file exists first
        if excel_service.file_exists():
            # Use Excel data
            player_df = excel_service.get_dataframe('PLAYERDETAILS')
            match_df = excel_service.get_dataframe('MATCHDETAILS')
            
            if player_df.empty:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found or empty'}), 404
            if match_df.empty:
                return jsonify({'error': 'MATCHDETAILS worksheet not found or empty'}), 404
            
            # Convert to records format for compatibility
            player_records = player_df.to_dict('records')
            match_records = match_df.to_dict('records')
        
        else:
            # Fallback to Google Sheets
            client = get_google_sheets_client('ahly_match')
            if not client:
                return jsonify({'error': 'No Excel file uploaded and Google Sheets client not available'}), 500

            spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
            
            # Get player details for goals and assists
            try:
                player_sheet = spreadsheet.worksheet('PLAYERDETAILS')
                player_records = player_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found'}), 404
            
            # Get match details for championship information
            try:
                match_sheet = spreadsheet.worksheet('MATCHDETAILS')
                match_records = match_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get lineup details for matches and minutes
        try:
            lineup_sheet = spreadsheet.worksheet('LINEUPDETAILS')
            lineup_records = lineup_sheet.get_all_records()
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'LINEUPDETAILS worksheet not found'}), 404
        
        # Find player's goals and assists
        player_ga_records = []
        for record in player_records:
            if record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                ga_type = record.get('GA', '').strip().upper()
                if ga_type in ['GOAL', 'ASSIST']:
                    player_ga_records.append(record)
        
        # Group by championship
        championship_stats = {}
        
        for ga_record in player_ga_records:
            match_id = ga_record.get('MATCH_ID', '')
            if not match_id:
                continue
                
            # Find match details
            match_info = None
            for match_record in match_records:
                if match_record.get('MATCH_ID', '').strip() == match_id:
                    match_info = match_record
                    break
            
            if not match_info:
                continue
                
            championship = match_info.get('CHAMPION', '').strip()
            if not championship:
                continue
                
            # Initialize championship if not exists
            if championship not in championship_stats:
                championship_stats[championship] = {
                    'CHAMPION': championship,
                    'matches': set(),
                    'goals': 0,
                    'assists': 0,
                    'minutes': 0
                }
            
            # Add to goals/assists
            ga_type = ga_record.get('GA', '').strip().upper()
            if ga_type == 'GOAL':
                championship_stats[championship]['goals'] += 1
            elif ga_type == 'ASSIST':
                championship_stats[championship]['assists'] += 1
            
            # Add match to set (to avoid duplicates)
            championship_stats[championship]['matches'].add(match_id)
        
        # Add minutes from lineup
        for lineup_record in lineup_records:
            if lineup_record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and lineup_record.get('TEAM', '').strip() != team_filter:
                    continue
                    
                match_id = lineup_record.get('MATCH_ID', '').strip()
                if not match_id:
                    continue
                    
                # Find match championship
                match_info = None
                for match_record in match_records:
                    if match_record.get('MATCH_ID', '').strip() == match_id:
                        match_info = match_record
                        break
                
                if not match_info:
                    continue
                    
                championship = match_info.get('CHAMPION', '').strip()
                if not championship:
                    continue
                    
                # Initialize championship if not exists
                if championship not in championship_stats:
                    championship_stats[championship] = {
                        'CHAMPION': championship,
                        'matches': set(),
                        'goals': 0,
                        'assists': 0,
                        'minutes': 0
                    }
                
                # Add match to set
                championship_stats[championship]['matches'].add(match_id)
                
                # Add minutes
                try:
                    minutes = int(lineup_record.get('MINTOTAL', 0) or 0)
                    championship_stats[championship]['minutes'] += minutes
                except (ValueError, TypeError):
                    pass
        
        # Convert to list and calculate match counts
        championships = []
        for champ_name, stats in championship_stats.items():
            stats['matches'] = len(stats['matches'])
            stats['ga_sum'] = stats['goals'] + stats['assists']
            championships.append(stats)
        
        # Sort by total G+A (descending)
        championships.sort(key=lambda x: x['ga_sum'], reverse=True)
        
        result = {
            'success': True,
            'player_name': player_name,
            'championships': championships
        }
        
        print(f"📊 Found {len(championships)} championships for '{player_name}'")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error loading player championships: {e}")
        return jsonify({'error': f'Failed to load player championships: {str(e)}'}), 500

@app.route('/api/player-seasons/<player_name>')
def api_player_seasons(player_name):
    """API endpoint to get player seasons from Google Sheets"""
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading seasons for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Check if Excel file exists first
        if excel_service.file_exists():
            # Use Excel data
            player_df = excel_service.get_dataframe('PLAYERDETAILS')
            match_df = excel_service.get_dataframe('MATCHDETAILS')
            
            if player_df.empty:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found or empty'}), 404
            if match_df.empty:
                return jsonify({'error': 'MATCHDETAILS worksheet not found or empty'}), 404
            
            # Convert to records format for compatibility
            player_records = player_df.to_dict('records')
            match_records = match_df.to_dict('records')
        
        else:
            # Fallback to Google Sheets
            client = get_google_sheets_client('ahly_match')
            if not client:
                return jsonify({'error': 'No Excel file uploaded and Google Sheets client not available'}), 500

            spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
            
            # Get player details for goals and assists
            try:
                player_sheet = spreadsheet.worksheet('PLAYERDETAILS')
                player_records = player_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found'}), 404
            
            # Get match details for season information
            try:
                match_sheet = spreadsheet.worksheet('MATCHDETAILS')
                match_records = match_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get lineup details for matches and minutes
        try:
            lineup_sheet = spreadsheet.worksheet('LINEUPDETAILS')
            lineup_records = lineup_sheet.get_all_records()
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'LINEUPDETAILS worksheet not found'}), 404
        
        # Find player's goals and assists
        player_ga_records = []
        for record in player_records:
            if record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                ga_type = record.get('GA', '').strip().upper()
                if ga_type in ['GOAL', 'ASSIST']:
                    player_ga_records.append(record)
        
        # Group by season
        season_stats = {}
        
        for ga_record in player_ga_records:
            match_id = ga_record.get('MATCH_ID', '')
            if not match_id:
                continue
                
            # Find match details
            match_info = None
            for match_record in match_records:
                if match_record.get('MATCH_ID', '').strip() == match_id:
                    match_info = match_record
                    break
            
            if not match_info:
                continue
                
            season = match_info.get('SEASON', '').strip()
            if not season:
                continue
                
            # Initialize season if not exists
            if season not in season_stats:
                season_stats[season] = {
                    'SEASON': season,
                    'matches': set(),
                    'goals': 0,
                    'assists': 0,
                    'minutes': 0
                }
            
            # Add to goals/assists
            ga_type = ga_record.get('GA', '').strip().upper()
            if ga_type == 'GOAL':
                season_stats[season]['goals'] += 1
            elif ga_type == 'ASSIST':
                season_stats[season]['assists'] += 1
            
            # Add match to set (to avoid duplicates)
            season_stats[season]['matches'].add(match_id)
        
        # Add minutes from lineup
        for lineup_record in lineup_records:
            if lineup_record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and lineup_record.get('TEAM', '').strip() != team_filter:
                    continue
                    
                match_id = lineup_record.get('MATCH_ID', '').strip()
                if not match_id:
                    continue
                    
                # Find match season
                match_info = None
                for match_record in match_records:
                    if match_record.get('MATCH_ID', '').strip() == match_id:
                        match_info = match_record
                        break
                
                if not match_info:
                    continue
                    
                season = match_info.get('SEASON', '').strip()
                if not season:
                    continue
                    
                # Initialize season if not exists
                if season not in season_stats:
                    season_stats[season] = {
                        'SEASON': season,
                        'matches': set(),
                        'goals': 0,
                        'assists': 0,
                        'minutes': 0
                    }
                
                # Add match to set
                season_stats[season]['matches'].add(match_id)
                
                # Add minutes
                try:
                    minutes = int(lineup_record.get('MINTOTAL', 0) or 0)
                    season_stats[season]['minutes'] += minutes
                except (ValueError, TypeError):
                    pass
        
        # Convert to list and calculate match counts
        seasons = []
        for season_name, stats in season_stats.items():
            stats['matches'] = len(stats['matches'])
            stats['ga_sum'] = stats['goals'] + stats['assists']
            seasons.append(stats)
        
        # Sort by season name (descending - newest first)
        seasons.sort(key=lambda x: x['SEASON'], reverse=True)
        
        result = {
            'success': True,
            'player_name': player_name,
            'seasons': seasons
        }
        
        print(f"📊 Found {len(seasons)} seasons for '{player_name}'")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error loading player seasons: {e}")
        return jsonify({'error': f'Failed to load player seasons: {str(e)}'}), 500

@app.route('/api/player-vs-teams/<player_name>')
def api_player_vs_teams(player_name):
    """API endpoint to get player vs teams from Google Sheets"""
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading vs teams for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Check if Excel file exists first
        if excel_service.file_exists():
            # Use Excel data
            player_df = excel_service.get_dataframe('PLAYERDETAILS')
            match_df = excel_service.get_dataframe('MATCHDETAILS')
            
            if player_df.empty:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found or empty'}), 404
            if match_df.empty:
                return jsonify({'error': 'MATCHDETAILS worksheet not found or empty'}), 404
            
            # Convert to records format for compatibility
            player_records = player_df.to_dict('records')
            match_records = match_df.to_dict('records')
        
        else:
            # Fallback to Google Sheets
            client = get_google_sheets_client('ahly_match')
            if not client:
                return jsonify({'error': 'No Excel file uploaded and Google Sheets client not available'}), 500

            spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
            
            # Get player details for goals and assists
            try:
                player_sheet = spreadsheet.worksheet('PLAYERDETAILS')
                player_records = player_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'PLAYERDETAILS worksheet not found'}), 404
            
            # Get match details for opponent team information
            try:
                match_sheet = spreadsheet.worksheet('MATCHDETAILS')
                match_records = match_sheet.get_all_records()
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get lineup details for matches and minutes
        try:
            lineup_sheet = spreadsheet.worksheet('LINEUPDETAILS')
            lineup_records = lineup_sheet.get_all_records()
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'LINEUPDETAILS worksheet not found'}), 404
        
        # Find player's goals and assists
        player_ga_records = []
        for record in player_records:
            if record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and record.get('TEAM', '').strip() != team_filter:
                    continue
                ga_type = record.get('GA', '').strip().upper()
                if ga_type in ['GOAL', 'ASSIST']:
                    player_ga_records.append(record)
        
        # Group by opponent team
        team_stats = {}
        
        for ga_record in player_ga_records:
            match_id = ga_record.get('MATCH_ID', '')
            if not match_id:
                continue
                
            # Find match details
            match_info = None
            for match_record in match_records:
                if match_record.get('MATCH_ID', '').strip() == match_id:
                    match_info = match_record
                    break
            
            if not match_info:
                continue
                
            opponent_team = match_info.get('OPPONENT TEAM', '').strip()
            if not opponent_team:
                continue
                
            # Initialize team if not exists
            if opponent_team not in team_stats:
                team_stats[opponent_team] = {
                    'OPPONENT_TEAM': opponent_team,
                    'matches': set(),
                    'goals': 0,
                    'assists': 0,
                    'minutes': 0
                }
            
            # Add to goals/assists
            ga_type = ga_record.get('GA', '').strip().upper()
            if ga_type == 'GOAL':
                team_stats[opponent_team]['goals'] += 1
            elif ga_type == 'ASSIST':
                team_stats[opponent_team]['assists'] += 1
            
            # Add match to set (to avoid duplicates)
            team_stats[opponent_team]['matches'].add(match_id)
        
        # Add minutes from lineup
        for lineup_record in lineup_records:
            if lineup_record.get('PLAYER NAME', '').strip() == player_name:
                if team_filter and lineup_record.get('TEAM', '').strip() != team_filter:
                    continue
                    
                match_id = lineup_record.get('MATCH_ID', '').strip()
                if not match_id:
                    continue
                    
                # Find match opponent team
                match_info = None
                for match_record in match_records:
                    if match_record.get('MATCH_ID', '').strip() == match_id:
                        match_info = match_record
                        break
                
                if not match_info:
                    continue
                    
                opponent_team = match_info.get('OPPONENT TEAM', '').strip()
                if not opponent_team:
                    continue
                    
                # Initialize team if not exists
                if opponent_team not in team_stats:
                    team_stats[opponent_team] = {
                        'OPPONENT_TEAM': opponent_team,
                        'matches': set(),
                        'goals': 0,
                        'assists': 0,
                        'minutes': 0
                    }
                
                # Add match to set
                team_stats[opponent_team]['matches'].add(match_id)
                
                # Add minutes
                try:
                    minutes = int(lineup_record.get('MINTOTAL', 0) or 0)
                    team_stats[opponent_team]['minutes'] += minutes
                except (ValueError, TypeError):
                    pass
        
        # Convert to list and calculate match counts
        vs_teams = []
        for team_name, stats in team_stats.items():
            stats['matches'] = len(stats['matches'])
            stats['ga_sum'] = stats['goals'] + stats['assists']
            vs_teams.append(stats)
        
        # Sort by total G+A (descending)
        vs_teams.sort(key=lambda x: x['ga_sum'], reverse=True)
        
        result = {
            'success': True,
            'player_name': player_name,
            'vs_teams': vs_teams
        }
        
        print(f"📊 Found {len(vs_teams)} vs teams for '{player_name}'")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error loading player vs teams: {e}")
        return jsonify({'error': f'Failed to load player vs teams: {str(e)}'}), 500

@app.route('/api/player-vs-goalkeepers/<player_name>')
def api_player_vs_goalkeepers(player_name):
    """API endpoint to get player vs goalkeepers - only goalkeepers who conceded goals FROM this specific player"""
    team_filter = request.args.get('team', '').strip()
    try:
        # URL decode the player name
        from urllib.parse import unquote
        player_name = unquote(player_name)
        
        print(f"Loading vs goalkeepers for player: {player_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        # Check if Excel file exists first
        if excel_service.file_exists():
            # Use Excel data
            player_df = excel_service.get_dataframe('PLAYERDETAILS')
            gk_df = excel_service.get_dataframe('GKDETAILS')
            match_df = excel_service.get_dataframe('MATCHDETAILS')
            
            if player_df.empty or gk_df.empty or match_df.empty:
                return jsonify({'vs_goalkeepers': []})
            
            # Convert to records format for compatibility
            player_records = player_df.to_dict('records')
            gk_records = gk_df.to_dict('records')
            match_records = match_df.to_dict('records')
        else:
            # Fallback to Google Sheets
            client = get_google_sheets_client('ahly_match')
            if not client:
                return jsonify({'vs_goalkeepers': []})

            spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
            
            # Fetch required sheets
            try:
                player_ws = spreadsheet.worksheet('PLAYERDETAILS')
                gk_ws = spreadsheet.worksheet('GKDETAILS')
                match_ws = spreadsheet.worksheet('MATCHDETAILS')
                
                try:
                    player_records = player_ws.get_all_records()
                    gk_records = gk_ws.get_all_records()
                    match_records = match_ws.get_all_records()
                except IndexError:
                    player_records = []
                    gk_records = []
                    match_records = []
            except gspread.WorksheetNotFound:
                return jsonify({'vs_goalkeepers': []})

        # Build match_id to opponent team mapping
        match_id_to_opponent = {}
        for match in match_records:
            match_id = match.get('MATCH_ID', '').strip()
            opponent = match.get('OPPONENT TEAM', '').strip()
            if match_id and opponent:
                match_id_to_opponent[match_id] = opponent

        # Find all goals scored by this player
        player_goals = {}
        for record in player_records:
            record_player = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            
            # Check if this is our target player
            if record_player != player_name:
                continue
                
            # Apply team filter if specified
            if team_filter and record_team != team_filter:
                continue
                
            match_id = record.get('MATCH_ID', '').strip()
            ga_type = record.get('GA', '').strip().upper()
            
            if match_id and ga_type == 'GOAL':
                if match_id not in player_goals:
                    player_goals[match_id] = 0
                player_goals[match_id] += 1

        if not player_goals:
            return jsonify({'vs_goalkeepers': []})

        # Find goalkeepers who were playing against this player when he scored
        vs_goalkeepers = []
        goalkeeper_stats = {}
        
        # Get all goals with their minutes for this player
        player_goals_with_minutes = {}
        for record in player_records:
            record_player = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            
            # Check if this is our target player
            if record_player != player_name:
                continue
                
            # Apply team filter if specified
            if team_filter and record_team != team_filter:
                continue
                
            match_id = record.get('MATCH_ID', '').strip()
            ga_type = record.get('GA', '').strip().upper()
            goal_minute = record.get('MINUTE', '').strip()
            
            if match_id and ga_type == 'GOAL':
                if match_id not in player_goals_with_minutes:
                    player_goals_with_minutes[match_id] = []
                
                # Convert minute to integer for comparison
                try:
                    minute_int = int(goal_minute) if goal_minute else 0
                except (ValueError, TypeError):
                    minute_int = 0
                    
                player_goals_with_minutes[match_id].append({
                    'minute': minute_int,
                    'ga_type': ga_type
                })
        
        # Group goalkeepers by match
        match_goalkeepers = {}
        for gk_record in gk_records:
            match_id = gk_record.get('MATCH_ID', '').strip()
            goalkeeper_name = gk_record.get('PLAYER NAME', '').strip()
            gk_team = gk_record.get('TEAM', '').strip()
            eleven_backup = gk_record.get('11/BAKEUP', '').strip().upper()
            submin = gk_record.get('SUBMIN', '').strip()
            
            if not goalkeeper_name or not match_id:
                continue
                
            # Get opponent team from match data
            opponent_team = match_id_to_opponent.get(match_id, '')
            
            # Only process opponent goalkeepers
            if gk_team != opponent_team:
                continue
                
            if match_id not in match_goalkeepers:
                match_goalkeepers[match_id] = []
            
            # Convert SUBMIN to integer
            try:
                submin_int = int(submin) if submin else 0
            except (ValueError, TypeError):
                submin_int = 0
            
            match_goalkeepers[match_id].append({
                'name': goalkeeper_name,
                'team': gk_team,
                'eleven_backup': eleven_backup,
                'submin': submin_int
            })
        
        # Process each match where the player scored
        for match_id, goals in player_goals_with_minutes.items():
            if match_id not in match_goalkeepers:
                continue
                
            opponent_gks = match_goalkeepers[match_id]
            
            # If only one goalkeeper, assign all goals to them
            if len(opponent_gks) == 1:
                gk_name = opponent_gks[0]['name']
                goals_count = len(goals)
                
                if gk_name not in goalkeeper_stats:
                    goalkeeper_stats[gk_name] = {
                        'GOALKEEPER_NAME': gk_name,
                        'goals': 0,
                        'pen_goals': 0
                    }
                goalkeeper_stats[gk_name]['goals'] += goals_count
                
            else:
                # Multiple goalkeepers - need to determine who was playing at each goal time
                # Sort goalkeepers by SUBMIN (substitution minute)
                opponent_gks.sort(key=lambda x: x['submin'])
                
                # Find starting goalkeeper (usually 11/BAKEUP = '11' or lowest SUBMIN)
                starting_gk = None
                for gk in opponent_gks:
                    if gk['eleven_backup'] == '11' or gk['submin'] == 0:
                        starting_gk = gk
                        break
                
                if not starting_gk and opponent_gks:
                    starting_gk = opponent_gks[0]  # Fallback to first goalkeeper
                
                # Process each goal
                for goal in goals:
                    goal_minute = goal['minute']
                    
                    # Determine which goalkeeper was playing at this goal minute
                    responsible_gk = starting_gk
                    
                    # Check if there was a substitution before this goal
                    for gk in opponent_gks:
                        if gk['submin'] > 0 and gk['submin'] <= goal_minute:
                            # This goalkeeper came in before the goal
                            responsible_gk = gk
                    
                    if responsible_gk:
                        gk_name = responsible_gk['name']
                        if gk_name not in goalkeeper_stats:
                            goalkeeper_stats[gk_name] = {
                                'GOALKEEPER_NAME': gk_name,
                                'goals': 0,
                                'pen_goals': 0
                            }
                        goalkeeper_stats[gk_name]['goals'] += 1

        # Convert to list and filter out goalkeepers with 0 goals
        vs_goalkeepers = [
            stats for stats in goalkeeper_stats.values() 
            if stats['goals'] > 0
        ]
        
        # Sort by goals (descending)
        vs_goalkeepers.sort(key=lambda x: x['goals'], reverse=True)
        
        print(f"Found {len(vs_goalkeepers)} goalkeepers who conceded goals from {player_name}")
        return jsonify({'vs_goalkeepers': vs_goalkeepers})
        
    except Exception as e:
        print(f"Error loading player vs goalkeepers: {e}")
        return jsonify({'error': f'Failed to load player vs goalkeepers: {str(e)}'}), 500

# Additional Google Sheets endpoints

@app.route('/api/OLD_player_matches/<player_name>')
def api_OLD_player_matches(player_name):
    """OLD Google Sheets endpoint - Return player's matches with DATE, SEASON, AHLY MANAGER, OPPONENT TEAM, goals, assists, minutes.
    Optional team filter via ?team=TEAM_NAME.
    """
    team_filter = request.args.get('team', '').strip()
    try:
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500

        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)

        # Load LINEUPDETAILS for minutes and to collect player match_ids
        try:
            lineup_ws = spreadsheet.worksheet('LINEUPDETAILS')
            try:
                lineup_records = lineup_ws.get_all_records()
            except IndexError:
                lineup_records = []
        except gspread.WorksheetNotFound:
            lineup_records = []

        # Aggregate minutes per match for the player (respect team filter if present)
        match_id_to_minutes = {}
        for rec in lineup_records:
            if rec.get('PLAYER NAME', '').strip() != player_name:
                continue
            if team_filter and rec.get('TEAM', '').strip() != team_filter:
                continue
            match_id = rec.get('MATCH_ID', '')
            minutes = rec.get('MINTOTAL', 0) or 0
            if match_id:
                match_id_to_minutes[match_id] = match_id_to_minutes.get(match_id, 0) + int(minutes)

        # Load PLAYERDETAILS for goals/assists per match (exact GA matches)
        try:
            player_ws = spreadsheet.worksheet('PLAYERDETAILS')
            try:
                player_records = player_ws.get_all_records()
            except IndexError:
                player_records = []
        except gspread.WorksheetNotFound:
            player_records = []

        match_id_to_ga = {}
        for rec in player_records:
            if rec.get('PLAYER NAME', '').strip() != player_name:
                continue
            if team_filter and rec.get('TEAM', '').strip() != team_filter:
                continue
            match_id = rec.get('MATCH_ID', '')
            if not match_id:
                continue
            ga_upper = str(rec.get('GA', '')).strip().upper()
            goals = 1 if ga_upper == 'GOAL' else 0
            assists = 1 if ga_upper == 'ASSIST' else 0
            if match_id not in match_id_to_ga:
                match_id_to_ga[match_id] = {'goals': 0, 'assists': 0}
            match_id_to_ga[match_id]['goals'] += goals
            match_id_to_ga[match_id]['assists'] += assists

        # If there are no matches for player, return empty list
        player_match_ids = set(match_id_to_minutes.keys()) | set(match_id_to_ga.keys())
        if not player_match_ids:
            return jsonify({'player_name': player_name, 'matches': []})

        # Load MATCHDETAILS once and index by MATCH_ID
        match_details_records = []
        try:
            match_ws = spreadsheet.worksheet('MATCHDETAILS')
            try:
                match_details_records = match_ws.get_all_records()
            except IndexError:
                match_details_records = []
        except gspread.WorksheetNotFound:
            # Try fallback name 'Match'
            try:
                match_ws = spreadsheet.worksheet('Match')
                match_details_records = match_ws.get_all_records()
            except gspread.WorksheetNotFound:
                match_details_records = []

        match_index = {}
        for rec in match_details_records:
            mid = rec.get('MATCH_ID') or rec.get('Match ID') or rec.get('match_id')
            if mid:
                match_index[str(mid)] = rec

        # Build result list for player's matches - only matches with goals or assists
        result = []
        for mid in player_match_ids:
            details = match_index.get(str(mid), {})
            goals = match_id_to_ga.get(mid, {}).get('goals', 0)
            assists = match_id_to_ga.get(mid, {}).get('assists', 0)
            
            # Only include matches where player scored goals or made assists
            if goals > 0 or assists > 0:
                result.append({
                    'MATCH_ID': mid,
                    'DATE': details.get('DATE') or details.get('Date') or details.get('date') or '',
                    'SEASON': details.get('SEASON', ''),
                    'AHLY MANAGER': details.get('AHLY MANAGER', ''),
                    'OPPONENT TEAM': details.get('OPPONENT TEAM', ''),
                    'goals': goals,
                    'assists': assists,
                    'minutes': match_id_to_minutes.get(mid, 0)
                })

        # Sort by DATE descending when possible
        def parse_date_simple(val):
            if not val or str(val).strip() == '':
                return datetime.min
            try:
                # Use Python's dateutil parser which handles many formats including "8-Mar-2024"
                from dateutil import parser
                return parser.parse(str(val).strip())
            except Exception:
                # Fallback to datetime.min for unparseable dates
                return datetime.min

        result.sort(key=lambda r: parse_date_simple(r['DATE']), reverse=True)

        # Add message if no matches with goals/assists found
        if not result:
            return jsonify({
                'player_name': player_name, 
                'matches': [],
                'message': 'No matches found where player scored goals or made assists'
            })

        return jsonify({'player_name': player_name, 'matches': result})

    except Exception as e:
        print(f"Error fetching player matches: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to fetch player matches: {str(e)}'}), 500
@app.route('/api/goalkeepers-data')
def api_goalkeepers_data():
    """Return all goalkeepers from PLAYERDATABASE with their teams"""
    try:
        # Check if Excel file exists first
        if excel_service.file_exists():
            # Use Excel data
            player_df = excel_service.get_dataframe('PLAYERDATABASE')
            if player_df.empty:
                return jsonify({'error': 'PLAYERDATABASE worksheet not found or empty'}), 404
            
            # Convert to records format for compatibility
            player_records = player_df.to_dict('records')
        
        else:
            # Fallback to Google Sheets
            client = get_google_sheets_client('ahly_match')
            if not client:
                return jsonify({'error': 'No Excel file uploaded and Google Sheets client not available'}), 500

            spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
            
            # Fetch PLAYERDATABASE
            try:
                player_ws = spreadsheet.worksheet('PLAYERDATABASE')
                try:
                    player_records = player_ws.get_all_records()
                except IndexError:
                    player_records = []
            except gspread.WorksheetNotFound:
                return jsonify({'error': 'PLAYERDATABASE worksheet not found'}), 404

        # Filter goalkeepers and extract unique data
        goalkeepers = []
        seen_goalkeepers = set()
        goalkeeper_teams = {}  # {player_name: [list of teams]}
        
        for record in player_records:
            player_name = record.get('PLAYER NAME', '').strip()
            team = record.get('TEAM', '').strip()
            position = record.get('POSITION', '').strip().upper()
            
            # Check if this is a goalkeeper
            if position in ['GK', 'GOALKEEPER', 'حارس مرمى', 'حارس'] and player_name:
                # Add goalkeeper to seen list
                if player_name not in seen_goalkeepers:
                    seen_goalkeepers.add(player_name)
                    goalkeeper_teams[player_name] = []
                
                # Add team to goalkeeper's teams list if not already there
                if team and team not in goalkeeper_teams[player_name]:
                    goalkeeper_teams[player_name].append(team)
        
        # Create goalkeepers list with all their teams
        for player_name in seen_goalkeepers:
            teams = goalkeeper_teams[player_name]
            goalkeepers.append({
                'name': player_name,
                'team': teams[0] if teams else '',  # First team as primary
                'teams': teams
            })

        return jsonify({'goalkeepers': goalkeepers})
        
    except Exception as e:
        print(f"Error fetching goalkeepers data: {e}")
        return jsonify({'error': f'Failed to fetch goalkeepers: {str(e)}'}), 500

@app.route('/api/goalkeeper-stats/<goalkeeper_name>')
def api_goalkeeper_stats(goalkeeper_name):
    """Return goalkeeper statistics for a specific goalkeeper"""
    try:
        # URL decode the goalkeeper name
        from urllib.parse import unquote
        goalkeeper_name = unquote(goalkeeper_name)
        team_filter = request.args.get('team', '')
        
        print(f"Loading goalkeeper stats for {goalkeeper_name.encode('utf-8', errors='ignore').decode('utf-8')} from API")
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500

        spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
        team_filter = request.args.get('team', '')
        
        # Fetch GKDETAILS sheet
        try:
            gk_ws = spreadsheet.worksheet('GKDETAILS')
            try:
                gk_records = gk_ws.get_all_records()
            except IndexError:
                gk_records = []
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'GKDETAILS worksheet not found'}), 404

        # Filter records for the specific goalkeeper
        filtered_records = []
        for record in gk_records:
            record_name = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            
            if record_name == goalkeeper_name:
                if not team_filter or record_team == team_filter:
                    filtered_records.append(record)

        # Convert records to goalkeeper stats format
        goalkeepers = []
        for record in filtered_records:
            goalkeepers.append({
                'name': record.get('PLAYER NAME', ''),
                'team': record.get('TEAM', ''),
                'matches': int(record.get('MATCHES', 0)),
                'cleanSheets': int(record.get('CLEAN SHEETS', 0)),
                'goalsConceded': int(record.get('GOALS CONCEDED', 0)),
                'saves': int(record.get('SAVES', 0))
            })

        result = {'goalkeepers': goalkeepers}
        
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error fetching goalkeeper stats: {e}")
        return jsonify({'error': f'Failed to fetch goalkeeper stats: {str(e)}'}), 500

@app.route('/api/gk-matches-apps-script/<goalkeeper_name>')
def api_gk_matches_apps_script(goalkeeper_name):
    """Return goalkeeper matches using Google Apps Script"""
    try:
        # URL decode the goalkeeper name
        from urllib.parse import unquote
        goalkeeper_name = unquote(goalkeeper_name)
        team_filter = request.args.get('team', '')
        print(f"Loading matches for goalkeeper using Apps Script: {goalkeeper_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        print(f"Loading matches for {goalkeeper_name.encode('utf-8', errors='ignore').decode('utf-8')} from Apps Script")
        
        # Call Google Apps Script for goalkeeper matches
        try:
            # Use the existing Google Apps Script URL from config
            apps_script_url = app.config.get('GOOGLE_APPS_SCRIPT_URL')
            if not apps_script_url:
                return jsonify({'error': 'Google Apps Script URL not configured'}), 500
            
            # Prepare data for Apps Script
            data = {
                'action': 'get_goalkeeper_matches',
            'goalkeeper_name': goalkeeper_name,
                'team_filter': team_filter
            }
            
            # Make request to Google Apps Script
            response = requests.post(apps_script_url, json=data, timeout=30)
            response.raise_for_status()
            
            result = response.json()
            
            if not result.get('success', False):
                return jsonify({'error': result.get('message', 'Unknown error from Apps Script')}), 500
                
        except requests.exceptions.Timeout:
            return jsonify({'error': 'Request timeout - Apps Script took too long to respond'}), 408
        except requests.exceptions.RequestException as e:
            return jsonify({'error': f'Error calling Apps Script: {str(e)}'}), 500
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid response from Apps Script'}), 500
        except Exception as e:
            return jsonify({'error': f'Unexpected error: {str(e)}'}), 500
        
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error fetching goalkeeper matches via Apps Script: {e}")
        return jsonify({'error': f'Failed to fetch goalkeeper matches: {str(e)}'}), 500

@app.route('/api/gk-matches/<goalkeeper_name>')
def api_gk_matches(goalkeeper_name):
    """Return goalkeeper matches for a specific goalkeeper"""
    try:
        # URL decode the goalkeeper name
        from urllib.parse import unquote
        goalkeeper_name = unquote(goalkeeper_name)
        team_filter = request.args.get('team', '')
        print(f"Loading matches for goalkeeper: {goalkeeper_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        print(f"Loading matches for {goalkeeper_name.encode('utf-8', errors='ignore').decode('utf-8')} from API")
        client = get_google_sheets_client('ahly_match')
        if not client:
            print("Google Sheets client not available")
            return jsonify({'error': 'Google Sheets client not available'}), 500

        spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
        print(f"Team filter: {team_filter}")
        
        # Get match IDs for this goalkeeper from GKDETAILS
        goalkeeper_match_ids = set()
        try:
            gk_ws = spreadsheet.worksheet('GKDETAILS')
            try:
                gk_records = gk_ws.get_all_records()
            except IndexError:
                gk_records = []
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'GKDETAILS worksheet not found'}), 404
        
        for record in gk_records:
            record_name = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            match_id = record.get('MATCH_ID', '')
            
            if record_name == goalkeeper_name:
                if not team_filter or record_team == team_filter:
                    if match_id:
                        goalkeeper_match_ids.add(match_id)
                        print(f"Found match ID: {match_id} for {goalkeeper_name.encode('utf-8', errors='ignore').decode('utf-8')}")
        
        print(f"Total match IDs found: {len(goalkeeper_match_ids)}")

        # Fetch MATCHDETAILS sheet
        try:
            match_ws = spreadsheet.worksheet('MATCHDETAILS')
            try:
                match_records = match_ws.get_all_records()
            except IndexError:
                match_records = []
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404

        # Filter matches for the specific goalkeeper
        matches = []
        for record in match_records:
            match_id = record.get('MATCH_ID', '')
            
            if match_id in goalkeeper_match_ids:
                # Get penalty goals and saves for this match
                pen_goals = 0
                pen_saves = 0
                
                # Count penalty goals from PLAYERDETAILS for this match
                try:
                    player_ws = spreadsheet.worksheet('PLAYERDETAILS')
                    try:
                        player_records = player_ws.get_all_records()
                    except IndexError:
                        player_records = []
                except gspread.WorksheetNotFound:
                    player_records = []
                
                for player_record in player_records:
                    player_match_id = str(player_record.get('MATCH_ID', '')).strip()
                    player_ga = str(player_record.get('GA', '')).strip().upper()
                    player_type = str(player_record.get('TYPE', '')).strip().upper()
                    player_team = str(player_record.get('TEAM', '')).strip()
                    
                    if (player_match_id == match_id and
                        player_ga == 'GOAL' and
                        player_type == 'PENGOAL' and
                        player_team != record.get('TEAM', '')):  # Only count opponent goals
                        pen_goals += 1
                
                # Count penalty saves from HOWPENMISSED for this match
                try:
                    howpen_ws = spreadsheet.worksheet('HOWPENMISSED')
                    try:
                        howpen_records = howpen_ws.get_all_records()
                    except IndexError:
                        howpen_records = []
                except gspread.WorksheetNotFound:
                    howpen_records = []
                
                for howpen_record in howpen_records:
                    howpen_match_id = str(howpen_record.get('MATCH_ID', '')).strip()
                    howpen_name = str(howpen_record.get('PLAYER NAME', '')).strip()
                    
                    if (howpen_match_id == match_id and
                        howpen_name == goalkeeper_name):
                        pen_saves += 1
                
                matches.append({
                    'date': record.get('DATE', ''),
                    'season': record.get('SEASON', ''),
                    'ahly_manager': record.get('AHLY MANAGER', ''),
                    'opponent_team': record.get('OPPONENT', ''),
                    'goals_conceded': int(record.get('GOALS CONCEDED', 0)),
                    'penalty_goals': pen_goals,
                    'penalty_saves': pen_saves,
                    'clean_sheet': 'Yes' if int(record.get('GOALS CONCEDED', 0)) == 0 else 'No'
                })

        print(f"Returning {len(matches)} matches")
        result = {'matches': matches}
        
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error fetching goalkeeper matches: {e}")
        error_msg = str(e)
        
        # Check if it's a quota exceeded error
        if 'quota' in error_msg.lower() or '429' in error_msg or 'rate_limit' in error_msg.lower():
            return jsonify({'error': 'Google Sheets API quota exceeded. Please wait a moment and try again.'}), 429
        else:
            return jsonify({'error': f'Failed to fetch goalkeeper matches: {error_msg}'}), 500

@app.route('/api/gk-championships/<goalkeeper_name>')
def api_gk_championships(goalkeeper_name):
    """Return goalkeeper championships for a specific goalkeeper"""
    try:
        # URL decode the goalkeeper name
        from urllib.parse import unquote
        goalkeeper_name = unquote(goalkeeper_name)
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500

        spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
        team_filter = request.args.get('team', '')
        
        # Fetch GKDETAILS sheet
        try:
            gk_ws = spreadsheet.worksheet('GKDETAILS')
            try:
                gk_records = gk_ws.get_all_records()
            except IndexError:
                gk_records = []
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'GKDETAILS worksheet not found'}), 404

        # Filter and aggregate records by championship
        championships = {}
        for record in gk_records:
            record_name = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            championship = record.get('CHAMPIONSHIP', '').strip()
            
            if record_name == goalkeeper_name and championship:
                if not team_filter or record_team == team_filter:
                    if championship not in championships:
                        championships[championship] = {
                            'championship': championship,
                            'matches': 0,
                            'clean_sheets': 0,
                            'goals_conceded': 0,
                            'saves': 0,
                            'save_percentage': 0
                        }
                    
                    championships[championship]['matches'] += int(record.get('MATCHES', 0))
                    championships[championship]['clean_sheets'] += int(record.get('CLEAN SHEETS', 0))
                    championships[championship]['goals_conceded'] += int(record.get('GOALS CONCEDED', 0))
                    championships[championship]['saves'] += int(record.get('SAVES', 0))

        # Calculate save percentage for each championship
        for champ in championships.values():
            total_shots = champ['saves'] + champ['goals_conceded']
            if total_shots > 0:
                champ['save_percentage'] = round((champ['saves'] / total_shots) * 100, 1)

        return jsonify({'championships': list(championships.values())})
        
    except Exception as e:
        print(f"Error fetching goalkeeper championships: {e}")
        return jsonify({'error': f'Failed to fetch goalkeeper championships: {str(e)}'}), 500

@app.route('/api/gk-seasons/<goalkeeper_name>')
def api_gk_seasons(goalkeeper_name):
    """Return goalkeeper seasons for a specific goalkeeper"""
    try:
        # URL decode the goalkeeper name
        from urllib.parse import unquote
        goalkeeper_name = unquote(goalkeeper_name)
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500

        spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
        team_filter = request.args.get('team', '')
        
        # Fetch GKDETAILS sheet
        try:
            gk_ws = spreadsheet.worksheet('GKDETAILS')
            try:
                gk_records = gk_ws.get_all_records()
            except IndexError:
                gk_records = []
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'GKDETAILS worksheet not found'}), 404

        # Filter and aggregate records by season
        seasons = {}
        for record in gk_records:
            record_name = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            season = record.get('SEASON', '').strip()
            
            if record_name == goalkeeper_name and season:
                if not team_filter or record_team == team_filter:
                    if season not in seasons:
                        seasons[season] = {
                            'season': season,
                            'matches': 0,
                            'clean_sheets': 0,
                            'goals_conceded': 0,
                            'saves': 0,
                            'save_percentage': 0
                        }
                    
                    seasons[season]['matches'] += int(record.get('MATCHES', 0))
                    seasons[season]['clean_sheets'] += int(record.get('CLEAN SHEETS', 0))
                    seasons[season]['goals_conceded'] += int(record.get('GOALS CONCEDED', 0))
                    seasons[season]['saves'] += int(record.get('SAVES', 0))

        # Calculate save percentage for each season
        for season in seasons.values():
            total_shots = season['saves'] + season['goals_conceded']
            if total_shots > 0:
                season['save_percentage'] = round((season['saves'] / total_shots) * 100, 1)

        return jsonify({'seasons': list(seasons.values())})
        
    except Exception as e:
        print(f"Error fetching goalkeeper seasons: {e}")
        return jsonify({'error': f'Failed to fetch goalkeeper seasons: {str(e)}'}), 500

@app.route('/api/gk-vs-teams/<goalkeeper_name>')
def api_gk_vs_teams(goalkeeper_name):
    """Return goalkeeper vs teams statistics for a specific goalkeeper"""
    try:
        # URL decode the goalkeeper name
        from urllib.parse import unquote
        goalkeeper_name = unquote(goalkeeper_name)
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500

        spreadsheet = client.open_by_key(app.config['SHEET_IDS']['ahly_match'])
        team_filter = request.args.get('team', '')
        
        # Fetch AHLY MATCH sheet
        try:
            match_ws = spreadsheet.worksheet('AHLY MATCH')
            try:
                match_records = match_ws.get_all_records()
            except IndexError:
                match_records = []
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'AHLY MATCH worksheet not found'}), 404

        # Filter and aggregate records by opponent team
        vs_teams = {}
        for record in match_records:
            record_name = record.get('PLAYER NAME', '').strip()
            record_team = record.get('TEAM', '').strip()
            opponent = record.get('OPPONENT', '').strip()
            
            if record_name == goalkeeper_name and opponent:
                if not team_filter or record_team == team_filter:
                    if opponent not in vs_teams:
                        vs_teams[opponent] = {
                            'team': opponent,
                            'matches': 0,
                            'clean_sheets': 0,
                            'goals_conceded': 0,
                            'saves': 0,
                            'save_percentage': 0
                        }
                    
                    vs_teams[opponent]['matches'] += 1
                    goals_conceded = int(record.get('GOALS CONCEDED', 0))
                    saves = int(record.get('SAVES', 0))
                    
                    vs_teams[opponent]['goals_conceded'] += goals_conceded
                    vs_teams[opponent]['saves'] += saves
                    
                    if goals_conceded == 0:
                        vs_teams[opponent]['clean_sheets'] += 1

        # Calculate save percentage for each team
        for team in vs_teams.values():
            total_shots = team['saves'] + team['goals_conceded']
            if total_shots > 0:
                team['save_percentage'] = round((team['saves'] / total_shots) * 100, 1)

        return jsonify({'vs_teams': list(vs_teams.values())})
        
    except Exception as e:
        print(f"Error fetching goalkeeper vs teams: {e}")
        return jsonify({'error': f'Failed to fetch goalkeeper vs teams: {str(e)}'}), 500

@app.route('/api/gk-vs-players/<goalkeeper_name>')
def api_gk_vs_players(goalkeeper_name):
    """Return goalkeeper vs players statistics for a specific goalkeeper (disabled; Excel-only mode)."""
    return jsonify({'player_name': goalkeeper_name, 'vs_goalkeepers': []})

@app.route('/api/goalkeeper-overview-stats/<goalkeeper_name>')
def api_goalkeeper_overview_stats(goalkeeper_name):
    """Return goalkeeper overview statistics: matches, minutes, goals conceded, clean sheets, pen goals conceded, pen saves.
    Optional team filter via ?team=TEAM_NAME.
    Sources: GKDETAILS (matches, minutes, goals_conceded), HOWPENMISSED (pen saves).
    """
    team_filter = request.args.get('team', '').strip()
    try:
        client = get_google_sheets_client('ahly_match')
        if not client:
            return jsonify({'error': 'Google Sheets client not available'}), 500

        sheet_id = app.config['SHEET_IDS']['ahly_match']
        spreadsheet = client.open_by_key(sheet_id)

        # Fetch GKDETAILS (goalkeeper stats)
        try:
            gk_ws = spreadsheet.worksheet('GKDETAILS')
            try:
                gk_records = gk_ws.get_all_records()
            except IndexError:
                gk_records = []
        except gspread.WorksheetNotFound:
            gk_records = []

        # Fetch HOWPENMISSED (penalty saves)
        try:
            pen_ws = spreadsheet.worksheet('HOWPENMISSED')
            try:
                pen_records = pen_ws.get_all_records()
            except IndexError:
                pen_records = []
        except gspread.WorksheetNotFound:
            pen_records = []

        # Filter goalkeeper records
        gk_matches = 0
        goals_conceded = 0
        clean_sheets = 0
        pen_goals_conceded = 0

        # Track matches with multiple goalkeepers from same team
        matches_with_multiple_gks = set()

        # First pass: check for multiple GKs per team in each match
        for rec in gk_records:
            rec_gk = str(rec.get('PLAYER NAME', '')).strip()
            if rec_gk != goalkeeper_name:
                continue
            if team_filter:
                rec_team = str(rec.get('TEAM', '')).strip()
                if rec_team != team_filter:
                    continue
            
            mid = rec.get('MATCH_ID') or rec.get('Match ID') or rec.get('match_id')
            if mid is None or str(mid).strip() == '':
                continue
            
            mid_str = str(mid)
            gk_team = str(rec.get('TEAM', '')).strip()
            
            # Check for multiple GKs from same team in this match
            same_team_gks = 0
            for other_rec in gk_records:
                other_mid = other_rec.get('MATCH_ID') or other_rec.get('Match ID') or other_rec.get('match_id')
                other_team = str(other_rec.get('TEAM', '')).strip()
                if str(other_mid) == mid_str and other_team == gk_team:
                    same_team_gks += 1
            
            if same_team_gks > 1:
                matches_with_multiple_gks.add(mid_str)

        # Second pass: calculate stats
        for rec in gk_records:
            rec_gk = str(rec.get('PLAYER NAME', '')).strip()
            if rec_gk != goalkeeper_name:
                continue
            if team_filter:
                rec_team = str(rec.get('TEAM', '')).strip()
                if rec_team != team_filter:
                    continue
            
            mid = rec.get('MATCH_ID') or rec.get('Match ID') or rec.get('match_id')
            if mid is None or str(mid).strip() == '':
                continue
            
            mid_str = str(mid)
            
            # Count matches
            gk_matches += 1
            
            # Goals conceded
            goals_conceded_val = rec.get('GOALS CONCEDED') or rec.get('Goals Conceded') or rec.get('goals_conceded') or 0
            try:
                goals_conceded_int = int(str(goals_conceded_val))
                goals_conceded += goals_conceded_int
                
                # Clean sheet only if no goals conceded AND this is the only GK from their team
                if goals_conceded_int == 0 and mid_str not in matches_with_multiple_gks:
                    clean_sheets += 1
            except Exception:
                pass

        # Count penalty saves from HOWPENMISSED
        pen_saves = 0
        for rec in pen_records:
            rec_gk = str(rec.get('PLAYER NAME', '')).strip()
            if rec_gk == goalkeeper_name:
                pen_saves += 1

        stats = {
            'total_matches': gk_matches,
            'goals_conceded': goals_conceded,
            'clean_sheets': clean_sheets,
            'pen_goals_conceded': pen_goals_conceded,  # Will be calculated separately
            'pen_saves': pen_saves
        }

        return jsonify({'goalkeeper_name': goalkeeper_name, 'stats': stats})

    except Exception as e:
        print(f"Error fetching goalkeeper overview stats: {e}")
        return jsonify({'error': f'Failed to fetch goalkeeper overview stats: {str(e)}'}), 500

@app.route('/api/finals-data')
def api_finals_data():
    """API endpoint to get Finals data from Google Sheets"""
    try:
        print("🏆 Loading Finals data from Google Sheets...")
        
        # Use ahlymatch credentials (or create specific finals credentials)
        creds_file = get_resource_path('credentials/ahlymatch.json')
        
        if not os.path.exists(creds_file):
            return jsonify({'error': 'Finals credentials file not found'}), 500
        
        # Initialize credentials and client
        creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        client = gspread.authorize(creds)
        
        # Open the Finals spreadsheet (update with actual sheet ID)
        sheet_id = 'YOUR_FINALS_SHEET_ID_HERE'  # Update this
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get Finals worksheet
        try:
            worksheet = spreadsheet.worksheet('Finals')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'Finals worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        filtered_records = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                filtered_records.append(cleaned_record)
        
        print(f"✅ Successfully loaded {len(filtered_records)} Finals records")
        
        return jsonify({
            'success': True,
            'records': filtered_records,
            'total_records': len(filtered_records)
        })
        
    except gspread.SpreadsheetNotFound:
        return jsonify({'error': 'Finals spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading Finals data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/pks-stats-data')
def api_pks_stats_data():
    """API endpoint to get PKS Stats data from Google Sheets"""
    try:
        # Check if force refresh is requested
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'
        
        # Try cache first (permanent cache - no expiration) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('pks_stats_data', ttl_hours=None)
            if cached_data:
                print(f"✅ Returning cached PKS stats data (permanent cache)")
                return jsonify(cached_data)
        else:
            print("🔄 Force refresh requested - bypassing cache")
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
        
        print("📊 Loading PKS Stats data from Google Sheets...")
        
        # Try to use environment variable first (for Render deployment)
        creds_json = app.config.get('GOOGLE_CREDENTIALS_JSON_AHLY_PKS')
        if creds_json:
            print("Using environment variable for PKS credentials")
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlypks.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'PKS credentials not found (neither env var nor file)'}), 500
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Open the PKS spreadsheet
        sheet_id = '1NM06fKzqEQc-K9XLgaIgd0PyQQAMHmOCVBKttQicZwY'
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PKS worksheet
        try:
            worksheet = spreadsheet.worksheet('PKS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'PKS worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        filtered_records = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                filtered_records.append(cleaned_record)
        
        print(f"✅ Successfully loaded {len(filtered_records)} PKS records")
        
        # Cache the result (permanent)
        result = {
            'success': True,
            'records': filtered_records,
            'total_records': len(filtered_records)
        }
        cache.set('pks_stats_data', result)
        
        return jsonify(result)
        
    except gspread.SpreadsheetNotFound:
        print("❌ PKS Spreadsheet not found")
        return jsonify({'error': 'PKS spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading PKS Stats data: {e}")
        return jsonify({'error': f'Failed to load PKS Stats data: {str(e)}'}), 500

@app.route('/api/finals-stats-data')
def api_finals_stats_data():
    """API endpoint to get Finals Stats data from Google Sheets"""
    try:
        # Check if force refresh is requested
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'
        
        # Try cache first (permanent cache - no expiration) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('finals_stats_data', ttl_hours=None)
            if cached_data:
                print(f"✅ Returning cached Finals stats data (permanent cache)")
                return jsonify(cached_data)
        else:
            print("🔄 Force refresh requested - bypassing cache")
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
        
        print("📊 Loading Finals Stats data from Google Sheets...")
        
        # Try to use environment variable first (for Render deployment)
        creds_json = app.config.get('GOOGLE_CREDENTIALS_JSON_AHLY_FINALS')
        if creds_json:
            print("Using environment variable for Finals credentials")
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyfinals.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Finals credentials not found (neither env var nor file)'}), 500
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Open the Finals spreadsheet
        sheet_id = '18lO8QMRqNUifGmFRZDTL58fwbb2k03HvkKyvzAq9HJc'
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get MATCHDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('MATCHDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        filtered_records = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                filtered_records.append(cleaned_record)
        
        print(f"✅ Successfully loaded {len(filtered_records)} Finals records")
        
        # Cache the result
        result = {
            'success': True,
            'records': filtered_records,
            'total_records': len(filtered_records)
        }
        cache.set('finals_stats_data', result)
        
        return jsonify(result)
        
    except gspread.SpreadsheetNotFound:
        print("❌ Finals Spreadsheet not found")
        return jsonify({'error': 'Finals spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading Finals Stats data: {e}")
        return jsonify({'error': f'Failed to load Finals Stats data: {str(e)}'}), 500

@app.route('/api/finals-players-data')
def api_finals_players_data():
    """API endpoint to get Finals Players data from PLAYERDETAILS sheet"""
    try:
        # Try cache first (permanent cache - no expiration)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('finals_players_data', ttl_hours=None)
        if cached_data:
            print(f"✅ Returning cached Finals players data (permanent cache)")
            return jsonify(cached_data)
        
        print("📊 Loading Finals Players data from Google Sheets...")
        
        # Try to use environment variable first (for Render deployment)
        creds_json = app.config.get('GOOGLE_CREDENTIALS_JSON_AHLY_FINALS')
        if creds_json:
            print("Using environment variable for Finals credentials")
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyfinals.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Finals credentials not found (neither env var nor file)'}), 500
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Open the Finals spreadsheet
        sheet_id = '18lO8QMRqNUifGmFRZDTL58fwbb2k03HvkKyvzAq9HJc'
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PLAYERDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'PLAYERDETAILS worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        filtered_records = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                filtered_records.append(cleaned_record)
        
        print(f"✅ Successfully loaded {len(filtered_records)} Finals Players records")
        
        # Cache the result (permanent)
        result = {
            'success': True,
            'records': filtered_records,
            'total_records': len(filtered_records)
        }
        cache.set('finals_players_data', result)
        
        return jsonify(result)
        
    except gspread.SpreadsheetNotFound:
        print("❌ Finals Spreadsheet not found")
        return jsonify({'error': 'Finals spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading Finals Players data: {e}")
        return jsonify({'error': f'Failed to load Finals Players data: {str(e)}'}), 500

@app.route('/api/finals-lineup-data')
def api_finals_lineup_data():
    """API endpoint to get Finals Lineup data from LINEUP11 sheet"""
    try:
        # Try cache first (permanent cache - no expiration)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('finals_lineup_data', ttl_hours=None)
        if cached_data:
            print(f"✅ Returning cached Finals lineup data (permanent cache)")
            return jsonify(cached_data)
        
        print("📊 Loading Finals Lineup data from Google Sheets...")
        
        # Try to use environment variable first (for Render deployment)
        creds_json = app.config.get('GOOGLE_CREDENTIALS_JSON_AHLY_FINALS')
        if creds_json:
            print("Using environment variable for Finals credentials")
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyfinals.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Finals credentials not found (neither env var nor file)'}), 500
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Open the Finals spreadsheet
        sheet_id = '18lO8QMRqNUifGmFRZDTL58fwbb2k03HvkKyvzAq9HJc'
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get LINEUP11 worksheet
        try:
            worksheet = spreadsheet.worksheet('LINEUP11')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'LINEUP11 worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        filtered_records = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                filtered_records.append(cleaned_record)
        
        print(f"✅ Successfully loaded {len(filtered_records)} Finals Lineup records")
        
        # Cache the result (permanent)
        result = {
            'success': True,
            'records': filtered_records,
            'total_records': len(filtered_records)
        }
        cache.set('finals_lineup_data', result)
        
        return jsonify(result)
        
    except gspread.SpreadsheetNotFound:
        print("❌ Finals Spreadsheet not found")
        return jsonify({'error': 'Finals spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading Finals Lineup data: {e}")
        return jsonify({'error': f'Failed to load Finals Lineup data: {str(e)}'}), 500

@app.route('/api/finals-playerdatabase-data')
def api_finals_playerdatabase_data():
    """API endpoint to get Player Database from PLAYERDATABASE sheet"""
    try:
        # Try cache first (permanent cache - no expiration)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('finals_playerdatabase_data', ttl_hours=None)
        if cached_data:
            print(f"✅ Returning cached Finals player database data (permanent cache)")
            return jsonify(cached_data)
        
        print("📊 Loading Finals Player Database from Google Sheets...")
        
        # Try to use environment variable first (for Render deployment)
        creds_json = app.config.get('GOOGLE_CREDENTIALS_JSON_AHLY_FINALS')
        if creds_json:
            print("Using environment variable for Finals credentials")
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyfinals.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Finals credentials not found (neither env var nor file)'}), 500
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Open the Finals spreadsheet
        sheet_id = '18lO8QMRqNUifGmFRZDTL58fwbb2k03HvkKyvzAq9HJc'
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PLAYERDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'PLAYERDATABASE worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        filtered_records = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                filtered_records.append(cleaned_record)
        
        print(f"✅ Successfully loaded {len(filtered_records)} Player Database records")
        
        # Cache the result (permanent)
        result = {
            'success': True,
            'records': filtered_records,
            'total_records': len(filtered_records)
        }
        cache.set('finals_playerdatabase_data', result)
        
        return jsonify(result)
        
    except gspread.SpreadsheetNotFound:
        print("❌ Finals Spreadsheet not found")
        return jsonify({'error': 'Finals spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading Finals Lineup data: {e}")
        return jsonify({'error': f'Failed to load Finals Lineup data: {str(e)}'}), 500

# ============================================================================
# CACHE MANAGEMENT API ENDPOINTS
# ============================================================================

@app.route('/api/cache/clear', methods=['POST'])
def api_clear_cache():
    """Clear backend cache (all or specific pattern)"""
    try:
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        
        data = request.get_json() or {}
        pattern = data.get('pattern')  # e.g., 'finals*' or 'pks*'
        
        cache.clear(pattern=pattern)
        
        message = f"Cache cleared" if not pattern else f"Cache cleared (pattern: {pattern})"
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        print(f"❌ Error clearing cache: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/refresh-cache')
def api_refresh_cache():
    """Refresh cache - clears all cached data to force reload from Google Sheets"""
    try:
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        
        # Clear all cache
        cache.clear()
        
        print("🔄 Cache cleared successfully - data will be reloaded from Google Sheets")
        return jsonify({
            'success': True, 
            'message': 'تم تحديث البيانات بنجاح! سيتم إعادة تحميل الصفحة...'
        })
        
    except Exception as e:
        print(f"❌ Error refreshing cache: {e}")
        return jsonify({
            'success': False, 
            'message': 'فشل في تحديث البيانات'
        }), 500

# ============================================================================
# GOOGLE SHEETS AUTO-SYNC API ENDPOINTS
# ============================================================================

@app.route('/api/ahly-stats/sheets-data', methods=['GET'])
def api_ahly_stats_sheets_data():
    """Get Al Ahly Stats data from Google Sheets (cached)"""
    try:
        from google_sheets_sync import get_sheets_data
        
        # Get data (from cache or sync if needed)
        data = get_sheets_data()
        
        if data:
            print(f"✅ Returning Al Ahly Stats data (sheets: {list(data.keys())})")
            return jsonify({
                'success': True,
                'data': data,
                'sheets': list(data.keys()),
                'timestamp': datetime.now().isoformat()
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to fetch data from Google Sheets'
            }), 500
            
    except Exception as e:
        print(f"❌ Error fetching Al Ahly Stats data: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ahly-stats/sync-now', methods=['POST'])
def api_ahly_stats_sync_now():
    """Manually trigger immediate sync from Google Sheets"""
    try:
        from google_sheets_sync import sync_now
        
        print("🔄 Manual sync triggered via API")
        
        # Perform sync
        success = sync_now()
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Sync completed successfully',
                'timestamp': datetime.now().isoformat()
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Sync failed'
            }), 500
            
    except Exception as e:
        print(f"❌ Error during manual sync: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ahly-stats/sync-status', methods=['GET'])
def api_ahly_stats_sync_status():
    """Get sync status information"""
    try:
        from google_sheets_sync import get_sync_status
        from scheduler_service import get_scheduler_status
        
        sync_status = get_sync_status()
        scheduler_status = get_scheduler_status()
        
        return jsonify({
            'success': True,
            'sync': sync_status,
            'scheduler': scheduler_status,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        print(f"❌ Error getting sync status: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================================
# AL AHLY VS ZAMALEK API ENDPOINTS
# ============================================================================

@app.route('/api/ahly-vs-zamalek/matches')
def api_ahly_vs_zamalek_matches():
    """API endpoint to get Al Ahly vs Zamalek matches data from Google Sheets"""
    print("🎯 API CALL: /api/ahly-vs-zamalek/matches")
    try:
        # Check if force refresh is requested
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'
        
        # Try cache first (6 hours TTL) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('ahly_vs_zamalek_matches', ttl_hours=6)
            if cached_data:
                print(f"✅ Returning cached data ({len(cached_data.get('matches', []))} matches)")
                return jsonify(cached_data)
        else:
            print("🔄 Force refresh requested - bypassing cache")
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
        
        print("⚽ Loading Al Ahly vs Zamalek matches data from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_AHLY_VS_ZAMALEK')
        if creds_env:
            print("Using environment variable for Al Ahly vs Zamalek credentials")
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyvszamalek.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Al Ahly vs Zamalek credentials not found'}), 500
            print(f"Using credentials file: {creds_file}")
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('AHLY_VS_ZAMALEK_SHEET_ID', '1jxRPyUQdqa38byIzorTfowbVUzL1pWLo2_KRLrvHN60')
        print(f"Using Sheet ID: {sheet_id}")
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get MATCHDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('MATCHDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'MATCHDETAILS worksheet not found'}), 404
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        matches = []
        for record in records:
            # Check if row has any non-empty values
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                # Convert all values to strings and clean them
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                matches.append(cleaned_record)
        
        print(f"✅ Loaded {len(matches)} matches from MATCHDETAILS")
        
        # Cache the result
        result = {'matches': matches, 'total': len(matches)}
        cache.set('ahly_vs_zamalek_matches', result)
        
        return jsonify(result)
        
    except gspread.SpreadsheetNotFound:
        print("❌ Al Ahly vs Zamalek Spreadsheet not found")
        return jsonify({'error': 'Spreadsheet not found'}), 404
    except Exception as e:
        print(f"❌ Error loading Al Ahly vs Zamalek matches: {e}")
        return jsonify({'error': f'Failed to load matches: {str(e)}'}), 500

@app.route('/api/ahly-vs-zamalek/player-details')
def api_ahly_vs_zamalek_player_details():
    """API endpoint to get Al Ahly vs Zamalek player details from PLAYERDETAILS sheet"""
    try:
        # Try cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('ahly_vs_zamalek_player_details', ttl_hours=6)
        if cached_data:
            print(f"✅ Returning cached Ahly vs Zamalek player details")
            return jsonify(cached_data)
        
        print("👥 Loading Al Ahly vs Zamalek player details from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_AHLY_VS_ZAMALEK')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyvszamalek.json')
            if not os.path.exists(creds_file):
                return jsonify({'playerDetails': []}), 200
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('AHLY_VS_ZAMALEK_SHEET_ID', '1jxRPyUQdqa38byIzorTfowbVUzL1pWLo2_KRLrvHN60')
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PLAYERDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'playerDetails': []}), 200
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        player_details = []
        for record in records:
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                player_details.append(cleaned_record)
        
        print(f"✅ Loaded {len(player_details)} player detail records")
        
        # Cache the result
        result = {'playerDetails': player_details}
        cache.set('ahly_vs_zamalek_player_details', result)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error loading player details: {e}")
        return jsonify({'playerDetails': []}), 200

@app.route('/api/ahly-vs-zamalek/lineupahly')
def api_ahly_vs_zamalek_lineup_ahly():
    """API endpoint to get Al Ahly lineup from LINEUPAHLY sheet"""
    try:
        # Try cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('ahly_vs_zamalek_lineup_ahly', ttl_hours=6)
        if cached_data:
            print(f"✅ Returning cached Ahly lineup")
            return jsonify(cached_data)
        
        print("📋 Loading Al Ahly lineup from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_AHLY_VS_ZAMALEK')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyvszamalek.json')
            if not os.path.exists(creds_file):
                return jsonify({'lineupAhly': []}), 200
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('AHLY_VS_ZAMALEK_SHEET_ID', '1jxRPyUQdqa38byIzorTfowbVUzL1pWLo2_KRLrvHN60')
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get LINEUPAHLY worksheet
        try:
            worksheet = spreadsheet.worksheet('LINEUPAHLY')
        except gspread.WorksheetNotFound:
            return jsonify({'lineupAhly': []}), 200
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        lineup = []
        for record in records:
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                lineup.append(cleaned_record)
        
        print(f"✅ Loaded {len(lineup)} Al Ahly lineup records")
        
        # Cache the result
        result = {'lineupAhly': lineup}
        cache.set('ahly_vs_zamalek_lineup_ahly', result)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error loading Al Ahly lineup: {e}")
        return jsonify({'lineupAhly': []}), 200

@app.route('/api/ahly-vs-zamalek/lineupzamalek')
def api_ahly_vs_zamalek_lineup_zamalek():
    """API endpoint to get Zamalek lineup from LINEUPZAMALEK sheet"""
    try:
        # Try cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('ahly_vs_zamalek_lineup_zamalek', ttl_hours=6)
        if cached_data:
            print(f"✅ Returning cached Zamalek lineup")
            return jsonify(cached_data)
        
        print("📋 Loading Zamalek lineup from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_AHLY_VS_ZAMALEK')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyvszamalek.json')
            if not os.path.exists(creds_file):
                return jsonify({'lineupZamalek': []}), 200
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('AHLY_VS_ZAMALEK_SHEET_ID', '1jxRPyUQdqa38byIzorTfowbVUzL1pWLo2_KRLrvHN60')
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get LINEUPZAMALEK worksheet
        try:
            worksheet = spreadsheet.worksheet('LINEUPZAMALEK')
        except gspread.WorksheetNotFound:
            return jsonify({'lineupZamalek': []}), 200
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        lineup = []
        for record in records:
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                lineup.append(cleaned_record)
        
        print(f"✅ Loaded {len(lineup)} Zamalek lineup records")
        
        # Cache the result
        result = {'lineupZamalek': lineup}
        cache.set('ahly_vs_zamalek_lineup_zamalek', result)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error loading Zamalek lineup: {e}")
        return jsonify({'lineupZamalek': []}), 200

@app.route('/api/ahly-vs-zamalek/player-database')
@app.route('/api/ahly-vs-zamalek/playerdatabase')
def api_ahly_vs_zamalek_player_database():
    """API endpoint to get player database from PLAYERDATABASE sheet"""
    try:
        # Try cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('ahly_vs_zamalek_player_database', ttl_hours=6)
        if cached_data:
            print(f"✅ Returning cached player database")
            return jsonify(cached_data)
        
        print("📊 Loading player database from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_AHLY_VS_ZAMALEK')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/alahlyvszamalek.json')
            if not os.path.exists(creds_file):
                return jsonify({'players': []}), 200
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('AHLY_VS_ZAMALEK_SHEET_ID', '1jxRPyUQdqa38byIzorTfowbVUzL1pWLo2_KRLrvHN60')
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PLAYERDATABASE worksheet
        try:
            worksheet = spreadsheet.worksheet('PLAYERDATABASE')
        except gspread.WorksheetNotFound:
            return jsonify({'players': []}), 200
        
        # Get all records as list of dictionaries
        records = worksheet.get_all_records()
        
        # Filter out empty rows
        player_database = []
        for record in records:
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                player_database.append(cleaned_record)
        
        print(f"✅ Loaded {len(player_database)} players from database")
        
        # Cache the result
        result = {'players': player_database}
        cache.set('ahly_vs_zamalek_player_database', result)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error loading player database: {e}")
        return jsonify({'players': []}), 200

@app.route('/api/egypt-teams/matches')
def api_egypt_teams_matches():
    """API endpoint to get Egypt National Teams matches from Google Sheets"""
    try:
        print("⚽ Loading Egypt National Teams matches from Google Sheets...")
        
        # Check if refresh is requested
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        # Try cache first (6 hours TTL) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('egypt_teams_matches', ttl_hours=6)
            if cached_data:
                print(f"✅ Returning cached Egypt Teams matches ({len(cached_data.get('matches', []))} matches)")
                return jsonify(cached_data)
        else:
            print("🔄 Force refresh requested - bypassing cache")
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_EGYPT_TEAMS')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/egyptnationalteam.json')
            print(f"Using credentials file: {creds_file}")
            if not os.path.exists(creds_file):
                print(f"❌ Credentials file not found: {creds_file}")
                return jsonify({'error': 'Credentials file not found', 'matches': []}), 404
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('EGYPT_TEAMS_SHEET_ID', '10PbAfoH9eqr4F82EBtO281RO42DgRzUzRv-dtELRDn8')
        print(f"Using Sheet ID: {sheet_id}")
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get MATCHDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('MATCHDETAILS')
        except gspread.WorksheetNotFound:
            print("❌ MATCHDETAILS worksheet not found")
            return jsonify({'error': 'No Data Available', 'matches': []}), 404
        
        # Get all records as list of dictionaries
        try:
            records = worksheet.get_all_records()
        except Exception as e:
            print(f"⚠️ MATCHDETAILS is empty or has no data: {e}")
            return jsonify({'error': 'No Data Available', 'matches': []}), 200
        
        # Filter out empty rows and clean data
        matches = []
        for record in records:
            # Check if row has any data
            has_data = any(str(v).strip() != '' for v in record.values())
            if has_data:
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = str(value).strip() if value else ''
                matches.append(cleaned_record)
        
        print(f"✅ Loaded {len(matches)} Egypt National Teams matches")
        
        # Cache the data
        result = {'matches': matches}
        cache.set('egypt_teams_matches', result)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error loading Egypt National Teams matches: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'matches': []}), 500

@app.route('/api/youth-egypt/matches', methods=['GET'])
def api_youth_egypt_matches():
    """API endpoint to get Youth Egypt Teams matches data"""
    try:
        # Check if refresh is requested
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        # Try cache first (6 hours TTL) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('youth_egypt_matches_data', ttl_hours=6)
            if cached_data:
                print(f"✅ Returning cached Youth Egypt Teams data ({len(cached_data)} records)")
                return jsonify({'success': True, 'records': cached_data})
        
        print("🇪🇬 Loading Youth Egypt Teams data from Google Sheets...")
        
        # Use Egypt Teams credentials (same sheet, different worksheets)
        client = get_google_sheets_client('egypt_match')
        if not client:
            return jsonify({'success': False, 'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for Youth Egypt
        sheet_id = app.config['SHEET_IDS']['youth_egypt']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get YouthMATCHDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('YouthMATCHDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'success': False, 'error': 'YouthMATCHDETAILS worksheet not found'}), 404
        
        # Get all records
        records = worksheet.get_all_records()
        
        
        # Clean data
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_records.append(cleaned_record)
        
        print(f"✅ Loaded {len(cleaned_records)} Youth Egypt Teams records from Google Sheets")
        
        # Cache the result for 6 hours
        cache.set('youth_egypt_matches_data', cleaned_records)
        
        return jsonify({'success': True, 'records': cleaned_records})
        
    except Exception as e:
        print(f"❌ Error loading Youth Egypt Teams data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/youth-egypt/players', methods=['GET'])
def api_youth_egypt_players():
    """API endpoint to get Youth Egypt Teams player details"""
    try:
        # Check if refresh is requested
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        # Try cache first (6 hours TTL) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('youth_egypt_players_data', ttl_hours=6)
            if cached_data:
                print(f"✅ Returning cached Youth Egypt Players data ({len(cached_data)} records)")
                return jsonify({'success': True, 'records': cached_data})
        
        print("👥 Loading Youth Egypt Players data from Google Sheets...")
        
        # Use Egypt Teams credentials (same sheet, different worksheets)
        client = get_google_sheets_client('egypt_match')
        if not client:
            return jsonify({'success': False, 'error': 'Google Sheets client not available'}), 500
        
        # Get the sheet ID for Youth Egypt
        sheet_id = app.config['SHEET_IDS']['youth_egypt']
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get YouthPLAYERDETAILS worksheet
        try:
            worksheet = spreadsheet.worksheet('YouthPLAYERDETAILS')
        except gspread.WorksheetNotFound:
            return jsonify({'success': False, 'error': 'YouthPLAYERDETAILS worksheet not found'}), 404
        
        # Get all records
        records = worksheet.get_all_records()
        
        
        # Clean data
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_records.append(cleaned_record)
        
        print(f"✅ Loaded {len(cleaned_records)} Youth Egypt Players records from Google Sheets")
        
        # Cache the result for 6 hours
        cache = get_cache_manager()
        cache.set('youth_egypt_players_data', cleaned_records)
        
        return jsonify({'success': True, 'records': cleaned_records})
        
    except Exception as e:
        print(f"❌ Error loading Youth Egypt Players data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/egypt-teams/players')
def api_egypt_teams_players():
    """API endpoint to get Egypt National Teams players with official goals"""
    try:
        print("👥 Loading Egypt National Teams players from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_EGYPT_TEAMS')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/egyptnationalteam.json')
            print(f"Using credentials file: {creds_file}")
            if not os.path.exists(creds_file):
                print(f"❌ Credentials file not found: {creds_file}")
                return jsonify({'error': 'Credentials file not found', 'players': []}), 404
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID from environment or use default
        sheet_id = os.environ.get('EGYPT_TEAMS_SHEET_ID', '10PbAfoH9eqr4F82EBtO281RO42DgRzUzRv-dtELRDn8')
        print(f"Using Sheet ID: {sheet_id}")
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PLAYERDATABASE worksheet
        try:
            player_db_worksheet = spreadsheet.worksheet('PLAYERDATABASE')
        except gspread.WorksheetNotFound:
            print("❌ PLAYERDATABASE worksheet not found")
            return jsonify({'error': 'No Data Available', 'players': []}), 404
        
        # Get PLAYERDETAILS worksheet
        try:
            player_details_worksheet = spreadsheet.worksheet('PLAYERDETAILS')
        except gspread.WorksheetNotFound:
            print("❌ PLAYERDETAILS worksheet not found")
            return jsonify({'error': 'No Data Available', 'players': []}), 404
        
        # Get MATCHDETAILS worksheet
        try:
            match_details_worksheet = spreadsheet.worksheet('MATCHDETAILS')
        except gspread.WorksheetNotFound:
            print("❌ MATCHDETAILS worksheet not found")
            return jsonify({'error': 'No Data Available', 'players': []}), 404
        
        # Get all records
        try:
            player_db_records = player_db_worksheet.get_all_records()
        except Exception as e:
            print(f"⚠️ PLAYERDATABASE is empty: {e}")
            return jsonify({'error': 'No Data Available', 'players': []}), 200
            
        try:
            player_details_records = player_details_worksheet.get_all_records()
        except Exception as e:
            print(f"⚠️ PLAYERDETAILS is empty: {e}")
            player_details_records = []
            
        try:
            match_details_records = match_details_worksheet.get_all_records()
        except Exception as e:
            print(f"⚠️ MATCHDETAILS is empty: {e}")
            match_details_records = []
        
        # Create a set of official match IDs (where CHAMPION SYSTEM = "OFI")
        official_match_ids = set()
        # Create a set of friendly match IDs (where CHAMPION SYSTEM = "FRI")
        friendly_match_ids = set()
        
        for match in match_details_records:
            champion_system = str(match.get('CHAMPION SYSTEM', '')).strip()
            match_id = str(match.get('MATCH_ID', '')).strip()
            if match_id:
                if champion_system == 'OFI':
                    official_match_ids.add(match_id)
                elif champion_system == 'FRI':
                    friendly_match_ids.add(match_id)
        
        print(f"Found {len(official_match_ids)} official matches and {len(friendly_match_ids)} friendly matches")
        
        # Calculate official and friendly goals, and official and friendly assists for each player
        players_official_goals = {}
        players_friendly_goals = {}
        players_official_assists = {}
        players_friendly_assists = {}
        
        for detail in player_details_records:
            player_name = str(detail.get('PLAYER NAME', '')).strip()
            if not player_name:
                continue
            
            ga_value = str(detail.get('GA', '')).strip()
            match_id = str(detail.get('MATCH_ID', '')).strip()
            
            gatotal = detail.get('GATOTAL', 0)
            try:
                gatotal = int(gatotal) if gatotal else 0
            except (ValueError, TypeError):
                gatotal = 0
            
            # Check if GA contains "GOAL" (exact match)
            if ga_value == 'GOAL':
                # Count official goals
                if match_id in official_match_ids:
                    if player_name in players_official_goals:
                        players_official_goals[player_name] += gatotal
                    else:
                        players_official_goals[player_name] = gatotal
                
                # Count friendly goals
                if match_id in friendly_match_ids:
                    if player_name in players_friendly_goals:
                        players_friendly_goals[player_name] += gatotal
                    else:
                        players_friendly_goals[player_name] = gatotal
            
            # Check if GA contains "ASSIST" (exact match)
            elif ga_value == 'ASSIST':
                # Count official assists
                if match_id in official_match_ids:
                    if player_name in players_official_assists:
                        players_official_assists[player_name] += gatotal
                    else:
                        players_official_assists[player_name] = gatotal
                
                # Count friendly assists
                if match_id in friendly_match_ids:
                    if player_name in players_friendly_assists:
                        players_friendly_assists[player_name] += gatotal
                    else:
                        players_friendly_assists[player_name] = gatotal
        
        # Get unique player names and their teams from PLAYERDATABASE
        player_teams = {}
        for player in player_db_records:
            name = str(player.get('PLAYER NAME', '')).strip()
            team = str(player.get('TEAM', '')).strip()
            if name:
                player_teams[name] = team
        
        # Build final players list
        players = []
        for player_name, team in player_teams.items():
            official_goals = players_official_goals.get(player_name, 0)
            friendly_goals = players_friendly_goals.get(player_name, 0)
            official_assists = players_official_assists.get(player_name, 0)
            friendly_assists = players_friendly_assists.get(player_name, 0)
            
            # Calculate total G+A
            total_ga = official_goals + friendly_goals + official_assists + friendly_assists
            
            players.append({
                'playerName': player_name,
                'team': team,
                'totalGA': total_ga,
                'officialGoals': official_goals,
                'friendlyGoals': friendly_goals,
                'officialAssists': official_assists,
                'friendlyAssists': friendly_assists
            })
        
        print(f"✅ Loaded {len(players)} players with goals and assists")
        return jsonify({'players': players})
        
    except Exception as e:
        print(f"❌ Error loading Egypt National Teams players: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'players': []}), 500

@app.route('/api/egypt-teams/player-details')
def api_egypt_teams_player_details():
    """API endpoint to get raw player details for client-side filtering"""
    try:
        print("👥 Loading Egypt National Teams player details...")
        
        # Check if refresh is requested
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        # Try cache first (6 hours TTL) unless force refresh
        if not force_refresh:
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
            cached_data = cache.get('egypt_teams_player_details', ttl_hours=6)
            if cached_data:
                print(f"✅ Returning cached Egypt Teams player details")
                return jsonify(cached_data)
        else:
            print("🔄 Force refresh requested - bypassing cache")
            from cache_manager import get_cache_manager
            cache = get_cache_manager()
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_EGYPT_TEAMS')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/egyptnationalteam.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Credentials file not found', 'playerDetails': [], 'playerDatabase': []}), 404
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID
        sheet_id = os.environ.get('EGYPT_TEAMS_SHEET_ID', '10PbAfoH9eqr4F82EBtO281RO42DgRzUzRv-dtELRDn8')
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get PLAYERDATABASE worksheet
        try:
            player_db_worksheet = spreadsheet.worksheet('PLAYERDATABASE')
        except:
            print("❌ PLAYERDATABASE worksheet not found")
            return jsonify({'error': 'No Data Available', 'playerDetails': [], 'playerDatabase': []}), 404
        
        # Get PLAYERDETAILS worksheet
        try:
            player_details_worksheet = spreadsheet.worksheet('PLAYERDETAILS')
        except:
            print("❌ PLAYERDETAILS worksheet not found")
            return jsonify({'error': 'No Data Available', 'playerDetails': [], 'playerDatabase': [], 'lineupDetails': []}), 404
        
        # Get LINEUPEGYPT worksheet
        try:
            lineup_egypt_worksheet = spreadsheet.worksheet('LINEUPEGYPT')
        except:
            print("❌ LINEUPEGYPT worksheet not found")
            return jsonify({'error': 'No Data Available', 'playerDetails': [], 'playerDatabase': [], 'lineupDetails': []}), 404
        
        # Get LINEUPOPPONENT worksheet
        try:
            lineup_opponent_worksheet = spreadsheet.worksheet('LINEUPOPPONENT')
        except:
            print("❌ LINEUPOPPONENT worksheet not found")
            return jsonify({'error': 'No Data Available', 'playerDetails': [], 'playerDatabase': [], 'lineupDetails': []}), 404
        
        # Get GKDETAILS worksheet (optional)
        try:
            gk_details_worksheet = spreadsheet.worksheet('GKDETAILS')
            gk_details_records = gk_details_worksheet.get_all_records()
        except:
            gk_details_records = []
        
        # Get HOWPENMISSED worksheet (optional)
        try:
            howpen_worksheet = spreadsheet.worksheet('HOWPENMISSED')
            howpen_records = howpen_worksheet.get_all_records()
        except:
            howpen_records = []
        
        # Get all records
        try:
            player_db_records = player_db_worksheet.get_all_records()
        except:
            print("⚠️ PLAYERDATABASE is empty")
            return jsonify({'error': 'No Data Available', 'playerDetails': [], 'playerDatabase': []}), 200
        
        try:
            player_details_records = player_details_worksheet.get_all_records()
        except:
            player_details_records = []
        
        # Get records from both lineup sheets
        try:
            lineup_egypt_records = lineup_egypt_worksheet.get_all_records()
        except:
            lineup_egypt_records = []
        
        try:
            lineup_opponent_records = lineup_opponent_worksheet.get_all_records()
        except:
            lineup_opponent_records = []
        
        # Add source team identifier to each record
        for record in lineup_egypt_records:
            record['SOURCE_TEAM'] = 'EGYPT'
        
        for record in lineup_opponent_records:
            record['SOURCE_TEAM'] = 'OPPONENT'
        
        # Combine both lineup records
        lineup_details_records = lineup_egypt_records + lineup_opponent_records
        
        # Clean data
        cleaned_player_db = []
        for record in player_db_records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_player_db.append(cleaned_record)
        
        cleaned_player_details = []
        for record in player_details_records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_player_details.append(cleaned_record)
        
        cleaned_lineup_details = []
        for record in lineup_details_records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_lineup_details.append(cleaned_record)
        
        cleaned_gk_details = []
        for record in gk_details_records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_gk_details.append(cleaned_record)
        
        cleaned_howpen = []
        for record in howpen_records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_howpen.append(cleaned_record)
        
        
        # Cache the data
        result = {
            'playerDatabase': cleaned_player_db,
            'playerDetails': cleaned_player_details,
            'lineupDetails': cleaned_lineup_details,
            'gkDetails': cleaned_gk_details,
            'howPenMissed': cleaned_howpen
        }
        cache.set('egypt_teams_player_details', result)
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        print(f"❌ Error loading player details: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e), 'playerDetails': [], 'playerDatabase': []}), 500

@app.route('/api/egypt-teams-pks')
def api_egypt_teams_pks():
    """API endpoint to get PKS data from ETPKS sheet"""
    try:
        # Try to get from cache first (6 hours TTL)
        from cache_manager import get_cache_manager
        cache = get_cache_manager()
        cached_data = cache.get('egypt_teams_pks_data', ttl_hours=6)
        if cached_data:
            print(f"✅ Returning cached Egypt Teams PKS data ({len(cached_data)} records)")
            return jsonify({'records': cached_data})
        
        print("🥅 Loading Egypt Teams PKS data from Google Sheets...")
        
        # Check environment variable first
        creds_env = os.environ.get('GOOGLE_CREDENTIALS_JSON_EGYPT_TEAMS')
        if creds_env:
            creds_info = json.loads(creds_env)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPE)
        else:
            # Fallback to local file
            creds_file = get_resource_path('credentials/egyptnationalteam.json')
            if not os.path.exists(creds_file):
                return jsonify({'error': 'Credentials file not found', 'records': []}), 404
            creds = Credentials.from_service_account_file(creds_file, scopes=SCOPE)
        
        client = gspread.authorize(creds)
        
        # Get Sheet ID (same as Egypt Teams)
        sheet_id = os.environ.get('EGYPT_TEAMS_SHEET_ID', '10PbAfoH9eqr4F82EBtO281RO42DgRzUzRv-dtELRDn8')
        spreadsheet = client.open_by_key(sheet_id)
        
        # Get ETPKS worksheet
        try:
            pks_worksheet = spreadsheet.worksheet('ETPKS')
        except gspread.WorksheetNotFound:
            return jsonify({'error': 'No Data Available', 'records': []}), 404
        
        # Get all records
        try:
            pks_records = pks_worksheet.get_all_records()
        except Exception as e:
            print(f"⚠️ ETPKS is empty: {e}")
            return jsonify({'error': 'No Data Available', 'records': []}), 200
        
        # Clean data
        cleaned_records = []
        for record in pks_records:
            cleaned_record = {}
            for key, value in record.items():
                cleaned_record[key] = str(value).strip() if value else ''
            cleaned_records.append(cleaned_record)
        
        print(f"✅ Loaded {len(cleaned_records)} PKS records from Google Sheets")
        
        # Cache the result for 6 hours
        cache.set('egypt_teams_pks_data', cleaned_records)
        
        return jsonify({'records': cleaned_records})
        
    except Exception as e:
        print(f"❌ Error loading PKS data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'records': []}), 500

if __name__ == '__main__':
    try:
        import webview
        import threading
        DESKTOP_MODE = True
    except ImportError:
        DESKTOP_MODE = False
    
    # Start Google Sheets Auto-Sync Scheduler
    try:
        from scheduler_service import start_scheduler
        print("\n" + "="*60)
        print("🚀 Starting Google Sheets Auto-Sync Scheduler")
        print("="*60)
        start_scheduler(sync_interval_hours=6)  # Sync every 6 hours
        print("✅ Scheduler started successfully")
        print("="*60 + "\n")
    except Exception as e:
        print(f"⚠️ Failed to start scheduler: {e}")
        print("   App will continue without auto-sync")
    
    if DESKTOP_MODE:
        # Function to run Flask server
        def run_server():
            app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False)
        
        # Start Flask server in a separate thread
        server_thread = threading.Thread(target=run_server, daemon=True)
        server_thread.start()
        
        # Create and start the webview window (desktop application)
        webview.create_window(
            'Football Database', 
            'http://127.0.0.1:5000',
            width=1400,
            height=900,
            resizable=True,
            fullscreen=False
        )
        webview.start()
    else:
        # Web deployment mode - just run Flask
        port = int(os.environ.get('PORT', 5000))
        app.run(host='0.0.0.0', port=port)
